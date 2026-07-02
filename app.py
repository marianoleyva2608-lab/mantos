import os, io, base64, sqlite3, json, hashlib
from flask import Flask, request, send_file, jsonify
from qr_catalog import CATEGORIA_FIJA as QR_CATEGORIA_FIJA, GRUPOS as QR_GRUPOS, PLANTA as QR_PLANTA, PROVEEDORES as QR_PROVEEDORES

app = Flask(__name__)
DATA_DIR = os.environ.get('DATA_DIR', '/app/data')
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH = os.path.join(DATA_DIR, 'reports.db')

def get_db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def init_db():
    con = get_db()
    con.execute('CREATE TABLE IF NOT EXISTS reports (id INTEGER PRIMARY KEY, machine_id TEXT, fecha TEXT, data TEXT)')
    con.execute('CREATE TABLE IF NOT EXISTS work_orders (id INTEGER PRIMARY KEY AUTOINCREMENT, numero TEXT NOT NULL, solicitante TEXT, fecha TEXT, equipo TEXT, planta TEXT, tipo TEXT, estatus TEXT, hora_inicio TEXT, hora_termino TEXT, tiempo_paro TEXT, descripcion_falla TEXT, actividad_realizada TEXT, refaccion TEXT, observaciones TEXT, firma_solicitante TEXT, firma_recibe TEXT, firma_liberacion TEXT, fotos TEXT, created_at TEXT DEFAULT (datetime(\'now\')))')
    # Migrate: add fotos column if missing
    try:
        con.execute('ALTER TABLE work_orders ADD COLUMN fotos TEXT DEFAULT "[]"')
        con.commit()
    except Exception:
        pass  # column already exists

    con.execute('''CREATE TABLE IF NOT EXISTS refacciones (
                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                   nombre TEXT, descripcion TEXT, marca TEXT, modelo TEXT,
                   categoria TEXT, criticidad TEXT DEFAULT 'MEDIA', seccion TEXT,
                   cant_min INTEGER DEFAULT 1, stock_actual INTEGER DEFAULT 0,
                   tiempo_entrega TEXT, proveedor TEXT, ubicacion TEXT,
                   costo REAL DEFAULT 0, notas TEXT, foto_b64 TEXT,
                   created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                   updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    # -- Seed refacciones AD-PACK (limpio, deduplicado, stock real fotos) --
    import base64 as _b64c, json as _jc
    _seed = _jc.loads(_b64c.b64decode("W1siUmVsw6kgdMOpcm1pY28gQ0hJTlQgTlIyLTI1IDIuNUEiLCAiUmVsZXZhZG9yIiwgIkFMVEEiLCAiRWxlY3RyaWNvIC8gQ29udHJvbCIsIDEsIDEsICIx4oCTMiBkw61hcyIsICJFTE1FU0kiLCAiRXN0YW50ZTogQ09OVEFDVE9SRVMiLCAzMjAuMCwgIiIsICJDSElOVCBOUjItM0UiLCAiMjA5MTA2Il0sIFsiU1NSIEpNUCBKTy0zNEYgOTBBIDQ4MFZBQyIsICJTU1IiLCAiQUxUQSIsICJFbGVjdHJpY28gLyBDb250cm9sIiwgMSwgMSwgIjPigJM1IGTDrWFzIiwgIkpNUCAvIENPSU5TQSIsICJFc3RhbnRlOiBIRVJSQU1JRU5UQSIsIDE4MDAuMCwgIkVxdWlwb3MgZ3JhbmRlcyIsICJKTVAgSkctMzRGIEQtNDgwQTUwWlMtTCIsICIiXSwgWyJSZWxldmFkb3IgRmluZGVyIDYwLjEyIDEwQSAyMzBWQUMgOC1waW4iLCAiUmVsZXZhZG9yIiwgIk1FRElBIiwgIkVsZWN0cmljbyAvIENvbnRyb2wiLCA0LCA0LCAiMeKAkzIgZMOtYXMiLCAiRUxNRVNJIC8gQ09JTlNBIiwgIkVzdGFudGU6IEVMRUNUUk8gVsOBTFZVTEFTIiwgMTgwLjAsICJWYXJpb3Mg4pyTIiwgIkZpbmRlciIsICJPbXJvbiJdLCBbIkZ1ZW50ZSAyNFZEQyBNb2VsbGVyIGVhc3k0MDAtUE9XIiwgIkZ1ZW50ZSBEQyIsICJBTFRBIiwgIkVsZWN0cmljbyAvIENvbnRyb2wiLCAxLCAxLCAiNeKAkzcgZMOtYXMiLCAiRUxNRVNJIiwgIkVzdGFudGU6IEhFUlJBTUlFTlRBIiwgMjIwMC4wLCAiIiwgIk1vZWxsZXIgZWFzeTQwMC1QT1ciLCAiIl0sIFsiTWljcm9zd2l0Y2ggbMOtbWl0ZSAxNUEgMTI1LTQ4MFYiLCAiRmluIGRlIGNhcnJlcmEiLCAiTUVESUEiLCAiRWxlY3RyaWNvIC8gQ29udHJvbCIsIDIsIDMsICIy4oCTMyBkw61hcyIsICJKTVAgLyBDT0lOU0EiLCAiRXN0YW50ZTogSEVSUkFNSUVOVEEiLCAzODAuMCwgIjMgdWRzIOKckyIsICJKTVAgTW9kZWxvIDExNi0yMDMiLCAiIl0sIFsiS2l0IHRlcm1pbmFsZXMgYWlzbGFkYXMgVm9sdGVjayA1NSBwaWV6YXMiLCAiVGVybWluYWwiLCAiQkFKQSIsICJFbGVjdHJpY28gLyBDb250cm9sIiwgMiwgMiwgIjEgZMOtYSIsICJUdWsgLyBPZmZpY2UgRGVwb3QiLCAiRXN0YW50ZTogQ09OVEFDVE9SRVMiLCA4NS4wLCAiIiwgIlZvbHRlYyIsICJUdWsiXSwgWyJSZXNpc3RlbmNpYSB0dWJ1bGFyIHRpcG8gVSAyMjBWQUMgI0JOLTEiLCAiUmVzaXN0ZW5jaWEiLCAiQUxUQSIsICJSZXNpc3RlbmNpYXMgLyBDYWxlZmFjY2lvbiIsIDUsIDIwLCAiNeKAkzEwIGTDrWFzIiwgIlRDUiIsICJFc3RhbnRlOiBSRVNJU1RFTkNJQVMiLCAzODAuMCwgIkNhamEgfjIwIHB6YXMg4pyTIiwgIlRDUiIsICJTQUwwMDAwMiJdLCBbIlJlc2lzdGVuY2lhIHBsYW5hIDY1MFcgdGVybW9wYXIgSyAyNXg2LjVjbSIsICJSZXNpc3RlbmNpYSIsICJBTFRBIiwgIlJlc2lzdGVuY2lhcyAvIENhbGVmYWNjaW9uIiwgMSwgMSwgIjXigJMxMCBkw61hcyIsICJUQ1IiLCAiRXN0YW50ZTogUkVTSVNURU5DSUFTIiwgNjUwLjAsICIiLCAiVENSIFNBTDAwMDAyIiwgIiJdLCBbIlJlc2lzdGVuY2lhIHBsYW5hIDYyNVcgdGVybW9wYXIgSyAyNXg2LjVjbSIsICJSZXNpc3RlbmNpYSIsICJBTFRBIiwgIlJlc2lzdGVuY2lhcyAvIENhbGVmYWNjaW9uIiwgMSwgMSwgIjXigJMxMCBkw61hcyIsICJUQ1IiLCAiRXN0YW50ZTogUkVTSVNURU5DSUFTIiwgNjIwLjAsICIiLCAiVENSIFNBTDAwMDAyIE5FR1UiLCAiIl0sIFsiUmVzaXN0ZW5jaWEgY2Vyw6FtaWNhIDM1MFcgdGVybW9wYXIgSyAxMng2LjVjbSIsICJSZXNpc3RlbmNpYSIsICJBTFRBIiwgIlJlc2lzdGVuY2lhcyAvIENhbGVmYWNjaW9uIiwgMSwgMSwgIjXigJMxMCBkw61hcyIsICJUQ1IiLCAiRXN0YW50ZTogUkVTSVNURU5DSUFTIiwgMzUwLjAsICIiLCAiVENSIiwgIiJdLCBbIlRlcm1pbmFsIGRlIGNlcsOhbWljYSBwYXJhIHJlc2lzdGVuY2lhIiwgIkNvbnN1bWlibGUiLCAiTUVESUEiLCAiUmVzaXN0ZW5jaWFzIC8gQ2FsZWZhY2Npb24iLCAxMCwgMTAsICIz4oCTNSBkw61hcyIsICJUQ1IgLyBDT0lOU0EiLCAiRXN0YW50ZTogVEVSTS4gQ0VSw4FNSUNBIiwgNDUuMCwgIiIsICJHZW7DqXJpY28iLCAiIl0sIFsiVGVybW9wYXIgdGlwbyBLIGRlIHJlcHVlc3RvIiwgIlNlbnNvciIsICJBTFRBIiwgIlJlc2lzdGVuY2lhcyAvIENhbGVmYWNjaW9uIiwgNCwgMiwgIjPigJM1IGTDrWFzIiwgIkNPSU5TQSAvIEF1dG9uaWNzIiwgIkVzdGFudGU6IElUTSBURU1QLiIsIDI4MC4wLCAiIiwgIkdlbsOpcmljbyIsICJBdXRvbmljcyJdLCBbIlbDoWx2dWxhIHNvbGVub2lkZSBERSBXSVQgMlctMjUtTkMtRS1WSTUgMVwiIE5DIiwgIkVsZWN0cm92w6FsdnVsYSIsICJBTFRBIiwgIk5ldW1hdGljbyAvIEhpZHJhdWxpY28iLCAyLCAxLCAiM+KAkzUgZMOtYXMiLCAiREUgV0lUIC8gQ09JTlNBIiwgIkVzdGFudGU6IEVMRUNUUk8gVsOBTFZVTEFTIiwgOTUwLjAsICJOQyDinJMiLCAiREUgV0lUIDJXLTI1LU5DLUUtVkk1IiwgIiJdLCBbIkVsZWN0cm92w6FsdnVsYSBuZXVtw6F0aWNhIDQvMiB2w61hcyAyNFZEQyIsICJFbGVjdHJvdsOhbHZ1bGEiLCAiQUxUQSIsICJOZXVtYXRpY28gLyBIaWRyYXVsaWNvIiwgMiwgMSwgIjPigJM1IGTDrWFzIiwgIlNNQyAvIENPSU5TQSIsICJFc3RhbnRlOiBFTEVDVFJPIFbDgUxWVUxBUyIsIDc4MC4wLCAiIiwgIlNwb3J0cm9uaWMiLCAiU01DIl0sIFsiTWFuw7NtZXRybyBnbGljZXJpbmEgMOKAkzEwIGJhciAxLzRcIiIsICJJbnN0cnVtZW50byIsICJCQUpBIiwgIk5ldW1hdGljbyAvIEhpZHJhdWxpY28iLCAyLCAyLCAiMuKAkzMgZMOtYXMiLCAiQ09JTlNBIiwgIkVzdGFudGU6IE1BTk9NRVRST1MiLCAyMjAuMCwgIuKckyIsICJXaWthIiwgIlZhcmlvcyJdLCBbIkNvbmRlbnNhZG9yIGRlIG1hcmNoYSBtb3RvciAxMTBWIDUwLzYwSHoiLCAiQ2FwYWNpdG9yIiwgIk1FRElBIiwgIk5ldW1hdGljbyAvIEhpZHJhdWxpY28iLCAzLCAzLCAiMuKAkzMgZMOtYXMiLCAiQ09JTlNBIC8gR3JhaW5nZXIiLCAiRXN0YW50ZTogRUxFQ1RSTyBWw4FMVlVMQVMiLCAxODAuMCwgIuKckyIsICJHZW7DqXJpY28iLCAiIl0sIFsiUm9kYW1pZW50byAvIGJhbGVybyBVUkIgUlVNIiwgIlJvZGFtaWVudG8iLCAiTUVESUEiLCAiTmV1bWF0aWNvIC8gSGlkcmF1bGljbyIsIDIsIDEsICIz4oCTNSBkw61hcyIsICJEaXN0cmlidWlkb3JhIFNLRiIsICJFc3RhbnRlOiBST0RBTUlFTlRPUyIsIDI1MC4wLCAiVmVyaWZpY2FyICMgYW50ZXMgZGUgcGVkaXIiLCAiVVJCIFJVTSIsICIiXSwgWyJGaXR0aW5nIG5ldW3DoXRpY28gcHVzaC10by1jb25uZWN0IHN1cnRpZG8iLCAiRml0dGluZyIsICJCQUpBIiwgIk5ldW1hdGljbyAvIEhpZHJhdWxpY28iLCAxMCwgMTAsICIx4oCTMiBkw61hcyIsICJTTUMgLyBDT0lOU0EiLCAiRXN0YW50ZTogRUxFQ1RSTyBWw4FMVlVMQVMiLCA0NS4wLCAi4pyTIiwgIlNNQyIsICJQYXJrZXIiXSwgWyJTb2xkYWR1cmEgZXN0YcOxbyBUcnVwZXIgNjAvNDAgMW1tIDQ1MGciLCAiQ29uc3VtaWJsZSIsICJCQUpBIiwgIkNvbnN1bWlibGVzIC8gSW5zdW1vcyIsIDEsIDIsICIxIGTDrWEiLCAiVHJ1cGVyIC8gZmVycmV0ZXLDrWEiLCAiRXN0YW50ZTogSEVSUkFNSUVOVEEiLCAxODAuMCwgIlZhcmlvcyByb2xsb3Mg4pyTIiwgIlRydXBlciA2MCIsICI0MCJdLCBbIkNpbnRhIGFpc2xhbnRlIFBWQyBhbWFyaWxsYSIsICJDb25zdW1pYmxlIiwgIkJBSkEiLCAiQ29uc3VtaWJsZXMgLyBJbnN1bW9zIiwgNCwgMywgIjEgZMOtYSIsICJPZmZpY2UgRGVwb3QiLCAiRXN0YW50ZTogQ09OVEFDVE9SRVMiLCAzNS4wLCAiIiwgIlZvbHRlYyIsICIzTSJdLCBbIkNpbnRhIGZpYnJhIGRlIHZpZHJpbyBhbHRhIHRlbXAuIiwgIkNvbnN1bWlibGUiLCAiTUVESUEiLCAiQ29uc3VtaWJsZXMgLyBJbnN1bW9zIiwgMiwgMSwgIjPigJM1IGTDrWFzIiwgIkNPSU5TQSIsICJFc3RhbnRlOiBSRVNJU1RFTkNJQVMiLCAxMjAuMCwgIlBhcmEgcmVzaXN0ZW5jaWFzIiwgIkdlbsOpcmljbyIsICIiXSwgWyJDaW50YSB0ZWZsw7NuIDEvMlwiIiwgIkNvbnN1bWlibGUiLCAiQkFKQSIsICJDb25zdW1pYmxlcyAvIEluc3Vtb3MiLCA1LCAzLCAiMSBkw61hIiwgIkZlcnJldGVyw61hIiwgIkVzdGFudGU6IEhFUlJBTUlFTlRBIiwgMTguMCwgIiIsICJHZW7DqXJpY28iLCAiIl0sIFsiUGVnYW1lbnRvIGluc3RhbnTDoW5lbyBzdXBlciBnbHVlIiwgIkNvbnN1bWlibGUiLCAiQkFKQSIsICJDb25zdW1pYmxlcyAvIEluc3Vtb3MiLCAyLCAyLCAiMSBkw61hIiwgIkZlcnJldGVyw61hIiwgIkVzdGFudGU6IENPTlRBQ1RPUkVTIiwgNDUuMCwgIuKckyIsICJMb2N0aXRlIiwgIkdlbsOpcmljbyJdLCBbIkx1YnJpY2FudGUgYW50aS1hZ2Fycm90YW1pZW50byBMQiA3NzEiLCAiTHVicmljYW50ZSIsICJCQUpBIiwgIkNvbnN1bWlibGVzIC8gSW5zdW1vcyIsIDEsIDEsICI14oCTNyBkw61hcyIsICJHcmFpbmdlciIsICJFc3RhbnRlOiBSRVNJU1RFTkNJQVMiLCAzODAuMCwgIkxhdGEg4pyTIiwgIkxCIDc3MSIsICJKZXQtTHViZSJdLCBbIkFjZWl0ZSByZWZyaWdlcmFudGUgV8O8cnRoIDUwMG1sIHNwcmF5IiwgIkx1YnJpY2FudGUiLCAiQkFKQSIsICJDb25zdW1pYmxlcyAvIEluc3Vtb3MiLCAyLCAxLCAiM+KAkzUgZMOtYXMiLCAiV8O8cnRoIE1YIiwgIkVzdGFudGU6IFJPREFNSUVOVE9TIiwgMjgwLjAsICIiLCAiV8O8cnRoIDUyMCIsICIiXSwgWyJDb21wdWVzdG8gdMOpcm1pY28gQXJjdGljIiwgIkNvbnN1bWlibGUiLCAiQkFKQSIsICJDb25zdW1pYmxlcyAvIEluc3Vtb3MiLCAxLCAxLCAiM+KAkzUgZMOtYXMiLCAiQW1hem9uIiwgIkVzdGFudGU6IEhFUlJBTUlFTlRBIiwgMTUwLjAsICIiLCAiQXJjdGljIiwgIiJdLCBbIkNhYmxlIGNvbmR1Y3RvciBWaWFrb24gLyBDb25kdW1leCAocm9sbG8pIiwgIkNhYmxlIiwgIk1FRElBIiwgIkNvbnN1bWlibGVzIC8gSW5zdW1vcyIsIDEsIDEsICIz4oCTNSBkw61hcyIsICJWaWFrb24gLyBFTE1FU0kiLCAiRXN0YW50ZTogUkVTSVNURU5DSUFTIiwgMTIwMC4wLCAiQ2FsLiBzZWfDum4gYXBsaWNhY2nDs24iLCAiVmlha29uIiwgIkNvbmR1bWV4Il0sIFsiQ29udGFjdG9yIENISU5UIE5YQy0zMiAzMkEgMjIwViIsICJDb250YWN0b3IiLCAiQUxUQSIsICJFbGVjdHJpY28gLyBDb250cm9sIiwgMiwgNCwgIjEtMiBkaWFzIiwgIkVMTUVTSSAvIENPSU5TQSIsICJFc3RhbnRlOiBDT05UQUNUT1JFUyIsIDAsICIiLCAiQ0hJTlQiLCAiTlhDLTMyIDkyNTIxMyJdLCBbIlJlbGV2YWRvciBBc2lhb24gMTBBIDI1MFZBQyIsICJSZWxldmFkb3IiLCAiTUVESUEiLCAiRWxlY3RyaWNvIC8gQ29udHJvbCIsIDIsIDEsICIxLTIgZGlhcyIsICJFTE1FU0kiLCAiRXN0YW50ZTogQ09OVEFDVE9SRVMiLCAwLCAiIiwgIkFzaWFvbiIsICI5MC4yLTEtMjIiXSwgWyJTU1IgR29sZCBTQVA0OTUwRCA1MEEiLCAiU1NSIiwgIkFMVEEiLCAiRWxlY3RyaWNvIC8gQ29udHJvbCIsIDEsIDEsICIzLTUgZGlhcyIsICJDT0lOU0EiLCAiRXN0YW50ZTogSEVSUkFNSUVOVEEiLCAwLCAiIiwgIkdvbGQiLCAiU0FQNDk1MEQiXSwgWyJQdWxzYWRvciBuZWdybyAyMm1tIE5DIiwgIlB1bHNhZG9yIiwgIkJBSkEiLCAiRWxlY3RyaWNvIC8gQ29udHJvbCIsIDIsIDQsICIxLTIgZGlhcyIsICJDT0lOU0EgLyBFTE1FU0kiLCAiRXN0YW50ZTogQ09OVEFDVE9SRVMiLCAwLCAiIiwgIlNjaG5laWRlciAvIFpCMiIsICJaQjItQkEzIl0sIFsiUHVsc2Fkb3IgYW1hcmlsbG8gMjJtbSBpbHVtaW5hZG8iLCAiUHVsc2Fkb3IiLCAiQkFKQSIsICJFbGVjdHJpY28gLyBDb250cm9sIiwgMSwgMSwgIjEtMiBkaWFzIiwgIkNPSU5TQSAvIEVMTUVTSSIsICJFc3RhbnRlOiBDT05UQUNUT1JFUyIsIDAsICIiLCAiU2NobmVpZGVyIC8gWkIyIiwgIlpCMi1CVzM1Il0sIFsiUHVsc2Fkb3IgYXp1bCAyMm1tIGlsdW1pbmFkbyIsICJQdWxzYWRvciIsICJCQUpBIiwgIkVsZWN0cmljbyAvIENvbnRyb2wiLCAxLCAxLCAiMS0yIGRpYXMiLCAiQ09JTlNBIC8gRUxNRVNJIiwgIkVzdGFudGU6IENPTlRBQ1RPUkVTIiwgMCwgIiIsICJTY2huZWlkZXIgLyBaQjIiLCAiWkIyLUJXMzYiXSwgWyJTZWxlY3RvciAyIHBvc2ljaW9uZXMgbGxhdmUgMjJtbSIsICJTZWxlY3RvciIsICJCQUpBIiwgIkVsZWN0cmljbyAvIENvbnRyb2wiLCAxLCAxLCAiMS0yIGRpYXMiLCAiQ09JTlNBIC8gRUxNRVNJIiwgIkVzdGFudGU6IENPTlRBQ1RPUkVTIiwgMCwgIiIsICJTY2huZWlkZXIgLyBaQjIiLCAiWkIyLUJHMiJdLCBbIlRlcm1vbWV0cm8gZGlnaXRhbCBUcmFjZWFibGUgYy9zb25kYSIsICJJbnN0cnVtZW50byIsICJNRURJQSIsICJFbGVjdHJpY28gLyBDb250cm9sIiwgMSwgMSwgIjUtNyBkaWFzIiwgIkZpc2hlciAvIEdyYWluZ2VyIiwgIkVzdGFudGU6IEhFUlJBTUlFTlRBIiwgMCwgIlBhcmEgdmVyaWZpY2FjaW9uIGRlIHRlbXBlcmF0dXJhIiwgIlRyYWNlYWJsZSIsICJUL0MiXSwgWyJDb250YWRvciBkaWdpdGFsIEF1dG9uaWNzIExFQk4iLCAiUmVzaXN0ZW5jaWEiLCAiQUxUQSIsICJSZXNpc3RlbmNpYXMgLyBDYWxlZmFjY2lvbiIsIDEsIDEsICI1LTEwIGRpYXMiLCAiQUVTQSAvIFRDUiIsICJFc3RhbnRlOiBSRVNJU1RFTkNJQVMiLCAwLCAiQ2VyYW1pYyBJbmZyYXJlZCBIZWF0ZXIiLCAiQUVTQSIsICIiXSwgWyJSZXNpc3RlbmNpYSBwbGFuYSAzMDBXIHRlcm1vcGFyIEsgMTJ4Ni41Y20iLCAiUmVzaXN0ZW5jaWEiLCAiQUxUQSIsICJSZXNpc3RlbmNpYXMgLyBDYWxlZmFjY2lvbiIsIDEsIDEsICI1LTEwIGRpYXMiLCAiVENSIiwgIkVzdGFudGU6IFJFU0lTVEVOQ0lBUyIsIDAsICIzIHBpZXphcyIsICJUQ1IiLCAiMjIwVkFDLzIzMCAzMjVXIl0sIFsiVmFsdnVsYSBzb2xlbm9pZGUgbmV1bWF0aWNhIDUvMiB2aWFzIG1hbmlmb2xkIiwgIkVsZWN0cm92YWx2dWxhIiwgIkFMVEEiLCAiTmV1bWF0aWNvIC8gSGlkcmF1bGljbyIsIDIsIDIsICIzLTUgZGlhcyIsICJTTUMgLyBDT0lOU0EiLCAiRXN0YW50ZTogRUxFQ1RSTyBWQUxWVUxBUyIsIDAsICJDb24gYm9iaW5hIDIyMFZBQyIsICJTTUMgLyBIVVlPIiwgIiJdLCBbIlbDoWx2dWxhIHNlZ3VyaWRhZCBLb21hdHN1IFNhZmUgVmFsdmUiLCAiVmFsdnVsYSIsICJBTFRBIiwgIk5ldW1hdGljbyAvIEhpZHJhdWxpY28iLCAxLCAxLCAiNy0xNCBkaWFzIiwgIktvbWF0c3UiLCAiRXN0YW50ZTogSEVSUkFNSUVOVEEiLCAwLCAiUGFydCBOby4gMDgwMTktMDA0OTgtMTAwIiwgIktvbWF0c3UiLCAiMDgwMTktMDA0OTgtMTAwIl0sIFsiTWFuw7NtZXRybyBBc2hjcm9mdCA2M21tIHZhY8OtbyIsICJJbnN0cnVtZW50byIsICJCQUpBIiwgIk5ldW1hdGljbyAvIEhpZHJhdWxpY28iLCAxLCAxLCAiMy01IGRpYXMiLCAiQ09JTlNBIC8gR3JhaW5nZXIiLCAiRXN0YW50ZTogTUFOT01FVFJPUyIsIDAsICI2MyAxMDA4IEEgMDJMIFhMSlpDIFZBQyIsICJBc2hjcm9mdCIsICI2MyAxMDA4Il0sIFsiTG9jdGl0ZSBMQiA3NzEgTmlja2VsIEFudGktU2VpemUgNDUzZyIsICJMdWJyaWNhbnRlIiwgIk1FRElBIiwgIkNvbnN1bWlibGVzIC8gSW5zdW1vcyIsIDEsIDEsICIzLTUgZGlhcyIsICJHcmFpbmdlciAvIEhlbmtlbCIsICJFc3RhbnRlOiBSRVNJU1RFTkNJQVMiLCAwLCAiQW50aWFnYXJyb3RhbWllbnRvIGhhc3RhIDEzMTVDIiwgIkxvY3RpdGUiLCAiTEIgNzcxIDEzNTU0MyJdLCBbIlNTUiBGb3RlayBTU1ItMjVWQSAyNUEiLCAiU1NSIiwgIkFMVEEiLCAiRWxlY3RyaWNvIC8gQ29udHJvbCIsIDIsIDEsICIzLTUgZGlhcyIsICJDT0lOU0EgLyBGb3RlayIsICJFc3RhbnRlOiBDT05UQUNUT1JFUyIsIDAsICJMb3RlIDIzM0YsIDMgdW5pZGFkZXMgZGlzcG9uaWJsZXMiLCAiRm90ZWsiLCAiU1NSLTI1VkEiXSwgWyJDb25kZW5zYWRvciBhcnJhbnF1ZS9tYXJjaGEgbW90b3IgMjIwVkFDIiwgIkNvbmRlbnNhZG9yIiwgIkFMVEEiLCAiRWxlY3RyaWNvIC8gQ29udHJvbCIsIDEsIDEsICIzLTUgZGlhcyIsICJFTE1FU0kgLyBHcmFpbmdlciIsICJFc3RhbnRlOiBIRVJSQU1JRU5UQSIsIDAsICJDYXBhY2l0b3IgZWxlY3Ryb2xpdGljbyBkZSBtb3RvciIsICJHZW5lcmljbyIsICIiXSwgWyJGdXNpYmxlcyBzdXJ0aWRvcyAoY2FqYSkiLCAiRnVzaWJsZSIsICJBTFRBIiwgIkVsZWN0cmljbyAvIENvbnRyb2wiLCA1LCAyMCwgIjEtMiBkaWFzIiwgIkVMTUVTSSAvIEZlcnJldGVyaWEiLCAiRXN0YW50ZTogSEVSUkFNSUVOVEEiLCAwLCAiRnVzaWJsZXMgZGUgdmlkcmlvIHkgY2VyYW1pY2Egc3VydGlkb3MiLCAiU3VydGlkbyIsICIiXSwgWyJDYWJsZSBNaWNhIFRhcGUgR2xhc3MgQnJhaWQgMTBBV0cgNTAwwrBDIDEwMG0iLCAiQ2FibGUiLCAiQUxUQSIsICJDb25zdW1pYmxlcyAvIEluc3Vtb3MiLCAxLCAyLCAiNS0xMCBkaWFzIiwgIlByb3ZlZWRvciBlc3BlY2lhbCIsICJFc3RhbnRlOiBURVJNSU5BTEVTIiwgMCwgIk1JQ0EvUFRGRSBUQVBFLCBQdXJlIE5pY2tlbCwgcm9sbG8gMTAwbSwgNC1KVU4tMjAyNSIsICJHZW5lcmljbyIsICIxMEFXRyA3NC8wLjMgNTAwQyJdLCBbIkNhYmxlIFZpYWtvbiBYWEkgUm9IUyAxNCBBV0cgbmVncm8gMTAwbSIsICJDYWJsZSIsICJNRURJQSIsICJDb25zdW1pYmxlcyAvIEluc3Vtb3MiLCAxLCAxLCAiMS0yIGRpYXMiLCAiVmlha29uIC8gQ29uZHVjdG9yZXMgTVRZIiwgIkVzdGFudGU6IFRFUk1JTkFMRVMiLCAwLCAiVEhXLUxTLCA5MEMsIHJvbGxvIDEwMG0iLCAiVmlha29uIiwgIlhYSSBSb0hTIDE0QVdHIl0sIFsiQ2FibGUgVmlha29uIDE2QVdHIG5lZ3JvIDEwMG0iLCAiQ2FibGUiLCAiTUVESUEiLCAiQ29uc3VtaWJsZXMgLyBJbnN1bW9zIiwgMSwgMSwgIjEtMiBkaWFzIiwgIlZpYWtvbiAvIENvbmR1Y3RvcmVzIE1UWSIsICJFc3RhbnRlOiBURVJNSU5BTEVTIiwgMCwgIjYwMFYgOTBDLCByb2xsbyAxMDBtLCBDYXQgMTI3MzA4IiwgIlZpYWtvbiIsICJURi1MUyAxNiAxMjczMDgiXSwgWyJDYWJsZSBWaWFrb24gMTIgQVdHIHJvam8gOTBDIDEwMG0iLCAiQ2FibGUiLCAiTUVESUEiLCAiQ29uc3VtaWJsZXMgLyBJbnN1bW9zIiwgMSwgMSwgIjEtMiBkaWFzIiwgIlZpYWtvbiAvIENvbmR1Y3RvcmVzIE1UWSIsICJFc3RhbnRlOiBURVJNSU5BTEVTIiwgMCwgIlRIVy1MUyBUSEhXLUxTLCA5MEMiLCAiVmlha29uIiwgIlhYSSAxMkFXRyByb2pvIl0sIFsiUm9kYW1pZW50byBVUkIgUGlsbG93IEJsb2NrIEJlYXJpbmciLCAiUm9kYW1pZW50byIsICJNRURJQSIsICJOZXVtYXRpY28gLyBIaWRyYXVsaWNvIiwgMSwgMSwgIjMtNSBkaWFzIiwgIkRpc3RyaWJ1aWRvcmEgTVRZIiwgIkVzdGFudGU6IFJPREFNSUVOVE9TIiwgMCwgIkV4cG9ydCB0byBHZXJtYW55LCBJU08gOTAwMSIsICJVUkIgLyBSVU0iLCAiUGlsbG93IEJsb2NrIl0sIFsiV8O8cnRoIFctTW90byBMdWJlIGdyYXNhIGNhZGVuYXMgMzAwbWwiLCAiTHVicmljYW50ZSIsICJCQUpBIiwgIkNvbnN1bWlibGVzIC8gSW5zdW1vcyIsIDEsIDEsICIzLTUgZGlhcyIsICJXdXJ0aCIsICJFc3RhbnRlOiBIRVJSQU1JRU5UQSIsIDAsICJMdWJyaWNhbnRlLWFudGljb3Jyb3Npdm8gcGFyYSBjYWRlbmFzIiwgIld1cnRoIiwgIlctTW90byBMdWJlIl0sIFsiV3VydGggQWNlaXRlIFJlZnJpZ2VyYW50ZSBjb3J0ZSA0MDBtbCIsICJMdWJyaWNhbnRlIiwgIkJBSkEiLCAiQ29uc3VtaWJsZXMgLyBJbnN1bW9zIiwgMSwgMSwgIjMtNSBkaWFzIiwgIld1cnRoIiwgIkVzdGFudGU6IEhFUlJBTUlFTlRBIiwgMCwgIkx1YnJpY2EsIHJlZnJpZ2VyYSB5IGNvbnNlcnZhIGhlcnJhbWllbnRhIGRlIGNvcnRlIiwgIld1cnRoIiwgIkFjZWl0ZSBSZWZyaWdlcmFudGUiXSwgWyJQYXN0YSBmbHV4IGRlIHNvbGRhZHVyYSBPbWVnYSAoZnJhc2NvIGFtYXJpbGxvKSIsICJDb25zdW1pYmxlIiwgIkJBSkEiLCAiQ29uc3VtaWJsZXMgLyBJbnN1bW9zIiwgMSwgMSwgIjEtMiBkaWFzIiwgIkZlcnJldGVyaWEgbG9jYWwiLCAiRXN0YW50ZTogSEVSUkFNSUVOVEEiLCAwLCAiRmx1eCBkZWNhcGFudGUgcGFyYSBzb2xkYWR1cmEgZWxlY3Ryb25pY2EiLCAiT21lZ2EiLCAiIl0sIFsiUGVnYW1lbnRvIGluc3RhbnTDoW5lbyBjaWFub2FjcmlsYXRvIiwgIkNvbnN1bWlibGUiLCAiQkFKQSIsICJDb25zdW1pYmxlcyAvIEluc3Vtb3MiLCAxLCAxLCAiMS0yIGRpYXMiLCAiRmVycmV0ZXJpYSBsb2NhbCIsICJCb2RlZ2EiLCAwLCAiU3VwZXIgYWRoZXNpdm8iLCAiR2VuZXJpY28iLCAiIl0sIFsiQ29uZWN0b3JlcyBwdXNoLXRvLWNvbm5lY3QgbmV1bcOhdGljb3Mgc3VydGlkbyIsICJDb25lY3RvciIsICJNRURJQSIsICJOZXVtYXRpY28gLyBIaWRyYXVsaWNvIiwgNSwgMTUsICIxLTMgZGlhcyIsICJDT0lOU0EgLyBGZXN0byAvIFNNQyIsICJFc3RhbnRlOiBFTEVDVFJPIFZBTFZVTEFTIiwgMCwgIlB1c2gtaW4gcGxhc3RpY28geSBsYXRvbiwgdmFyaWFzIG1lZGlkYXMgNi0xMm1tIiwgIlNNQyAvIFBhcmtlciIsICJQdXNoLWluIl0sIFsiVGVybWluYWxlcyBkZSBhcmdvbGxhIGRlc251ZGFzIHN1cnRpZG8gKGNhamEpIiwgIlRlcm1pbmFsIiwgIkJBSkEiLCAiQ29uc3VtaWJsZXMgLyBJbnN1bW9zIiwgNSwgMzAsICIxIGRpYSIsICJGZXJyZXRlcmlhIiwgIkVzdGFudGU6IFRFUk1JTkFMRVMiLCAwLCAiVGVybWluYWxlcyBkZSBjb2JyZSBzaW4gZm9ycm8gdmFyaW9zIGNhbGlicmVzIiwgIkdlbmVyaWNvIiwgIiJdLCBbIkJsb3F1ZXMgY2Vyw6FtaWNvcyB0ZXJtaW5hbGVzIDVUIiwgIlRlcm1pbmFsIiwgIk1FRElBIiwgIlJlc2lzdGVuY2lhcyAvIENhbGVmYWNjaW9uIiwgNSwgMTUsICIzLTUgZGlhcyIsICJUQ1IgLyBQcm92ZWVkb3IiLCAiRXN0YW50ZTogVEVSTUlOQUxFUyIsIDAsICJQYXJhIGNvbmV4aW9uIGRlIHJlc2lzdGVuY2lhcywgYWx0YSB0ZW1wZXJhdHVyYSIsICJHZW5lcmljbyIsICI1VCBjZXJhbWljIl0sIFsiQ2ludGEgZmlicmEgZGUgdmlkcmlvIGFpc2xhbnRlIChyb2xsbykiLCAiQ29uc3VtaWJsZSIsICJNRURJQSIsICJSZXNpc3RlbmNpYXMgLyBDYWxlZmFjY2lvbiIsIDEsIDEsICIzLTUgZGlhcyIsICJQcm92ZWVkb3IgZXNwZWNpYWwiLCAiRXN0YW50ZTogUkVTSVNURU5DSUFTIiwgMCwgIlBhcmEgYWlzbGFtaWVudG8gZGUgcmVzaXN0ZW5jaWFzIGFsdGEgdGVtcGVyYXR1cmEiLCAiR2VuZXJpY28iLCAiRWxlY3RyaWNhbCBmaWJlcmdsYXNzIl0sIFsiQ2ludGEgZmlicmEgZGUgdmlkcmlvIGFsdGEgdGVtcGVyYXR1cmEgKHJvbGxvKSIsICJFbMOpY3RyaWNvIiwgIkFsdGEiLCAiQWxtYWPDqW4iLCAxLCA1LCAiMy01IGTDrWFzIiwgIkF1dG9uaWNzIiwgIkVzdGFudGUtQXV0b25pY3MiLCAwLCAiVGVtcGVyYXR1cmUgbXVsdGktc2Vuc29yIHJlbGF5K1NTUiAxMDAtMjQwVkFDIiwgIkF1dG9uaWNzIiwgIlRDTjRTLTI0UiJdLCBbIlRpbWVyIEF1dG9uaWNzIEFUOE4gMTAwLTI0MFZBQyIsICJFbMOpY3RyaWNvIiwgIk1lZGlhIiwgIkFsbWFjw6luIiwgMiwgMiwgIjMtNSBkw61hcyIsICJBdXRvbmljcyIsICJFc3RhbnRlLUF1dG9uaWNzIiwgMCwgIlRpbWVyIG11bHRpZnVuY2nDs24gMTAwLTI0MFZBQy8yNC0yNDBWREMiLCAiQXV0b25pY3MiLCAiQVQ4TiJdLCBbIkNvbnRhY3RvciBGdWppIEVsZWN0cmljIFNDLTQtMSIsICJFbMOpY3RyaWNvIiwgIk1lZGlhIiwgIkFsbWFjw6luIiwgMSwgMSwgIjUtNyBkw61hcyIsICJGdWppIEVsZWN0cmljIiwgIkVzdGFudGUtQ29udGFjdG9yZXMiLCAwLCAiMjQwViA0a1cgLyA0NDBWIDcuNWtXIC8gNTUwViA5a1ciLCAiRnVqaSIsICJTQy00LTEiXSwgWyJJbnRlcnJ1cHRvciB0ZXJtb21hZ27DqXRpY28gTW9lbGxlciBYcG9sZSBDMiAyQSIsICJFbMOpY3RyaWNvIiwgIk1lZGlhIiwgIkFsbWFjw6luIiwgMSwgMSwgIjUtNyBkw61hcyIsICJNb2VsbGVyIiwgIkVzdGFudGUtQ29udGFjdG9yZXMiLCAwLCAiMSBwb2xvIDJBIGN1cnZhIEMgY2FycmlsIERJTiIsICJNb2VsbGVyIiwgIlBMU00tQzIiXSwgWyJJbnRlcnJ1cHRvciBhdXRvbcOhdGljbyBTaWVtZW5zIDJQIDEwQSIsICJFbMOpY3RyaWNvIiwgIk1lZGlhIiwgIkFsbWFjw6luIiwgMSwgMiwgIjUtNyBkw61hcyIsICJTaWVtZW5zIiwgIkVzdGFudGUtQ29udGFjdG9yZXMiLCAwLCAiMiBwb2xvcyBjYXJyaWwgRElOIiwgIlNpZW1lbnMiLCAiIl0sIFsiSW50ZXJydXB0b3IgYXV0b23DoXRpY28gU2NobmVpZGVyIGlDNjBOIiwgIkVsw6ljdHJpY28iLCAiTWVkaWEiLCAiQWxtYWPDqW4iLCAxLCAxLCAiNS03IGTDrWFzIiwgIlNjaG5laWRlciIsICJFc3RhbnRlLUNvbnRhY3RvcmVzIiwgMCwgImlDNjBOIGNhcnJpbCBESU4iLCAiU2NobmVpZGVyIiwgImlDNjBOIl0sIFsiUmVzaXN0ZW5jaWEgY2Vyw6FtaWNhIEFFU0EgMzI1VyIsICJSZWZhY2Npb25lcyIsICJBbHRhIiwgIkFsbWFjw6luIiwgMiwgMiwgIjUtNyBkw61hcyIsICJBRVNBMiIsICJFc3RhbnRlLVJlc2lzdGVuY2lhcyIsIDAsICJDZXJhbWljIEluZnJhcmVkIEhlYXRlciAyMjAvMjMwVkFDIDMyNVciLCAiQUVTQTIiLCAiIl0sIFsiUmVzaXN0ZW5jaWEgY2Vyw6FtaWNhIENlcmFtaWN4IDMyNVcgMTJ4Ni41Y20iLCAiUmVmYWNjaW9uZXMiLCAiQWx0YSIsICJBbG1hY8OpbiIsIDEsIDEsICI1LTcgZMOtYXMiLCAiQ2VyYW1pY3giLCAiRXN0YW50ZS1SZXNpc3RlbmNpYXMiLCAwLCAiSW5mcmFyZWQgZm9yIGluZHVzdHJ5IDIyMC8yMzBWQUMgMzI1VyAxMng2LjVjbSIsICJDZXJhbWljeCIsICIiXSwgWyJSZXNpc3RlbmNpYSBwbGFuYSAzMjVXIHRlcm1vcGFyIEsgMTJ4Ni41Y20iLCAiUmVmYWNjaW9uZXMiLCAiQWx0YSIsICJBbG1hY8OpbiIsIDIsIDQsICI1LTcgZMOtYXMiLCAiVENSIiwgIkVzdGFudGUtUmVzaXN0ZW5jaWFzIiwgMCwgIjIyMC8yMzBWQUMgMzI1VyBjb24gdGVybW9wYXIgSyAxMng2LjVjbSByZWYgU0FMMDAwMDIiLCAiVENSIiwgIlNBTDAwMDAyIl0sIFsiVsOhbHZ1bGEgc29sZW5vaWRlIERFIFdJVCAyVy0yNS1OQy1FLVZUNSIsICJOZXVtw6F0aWNvIiwgIkFsdGEiLCAiQWxtYWPDqW4iLCAxLCAxLCAiNS03IGTDrWFzIiwgIkRFIFdJVCIsICJFc3RhbnRlLUVsZWN0cm92YWx2dWxhcyIsIDAsICJTb2xlbm9pZGUgMiB2w61hcyBub3JtYWxtZW50ZSBjZXJyYWRhIDEgcHVsZ2FkYSIsICJERSBXSVQiLCAiMlctMjUtTkMtRS1WVDUiXSwgWyJUZXJtw7NtZXRybyBkaWdpdGFsIHBvcnTDoXRpbCBUcmFjZWFibGUiLCAiSW5zdHJ1bWVudGFjacOzbiIsICJNZWRpYSIsICJBbG1hY8OpbiIsIDEsIDEsICI1LTcgZMOtYXMiLCAiVHJhY2VhYmxlIiwgIkVzdGFudGUtSW5zdHJ1bWVudG9zIiwgMCwgIkNvbiBzb25kYSB0aXBvIGFndWphIGRpc3BsYXkgTENEIHRlcm1vcGFyIEsvSiIsICJUcmFjZWFibGUiLCAiIl0sIFsiTWFuw7NtZXRybyBhbmFsw7NnaWNvIG5ldW3DoXRpY28gY2lyY3VsYXIiLCAiSW5zdHJ1bWVudGFjacOzbiIsICJNZWRpYSIsICJBbG1hY8OpbiIsIDEsIDEsICI1LTcgZMOtYXMiLCAiR2Vuw6lyaWNvIiwgIkVzdGFudGUtTWFub21ldHJvcyIsIDAsICJNYW7Ds21ldHJvIGRlIHByZXNpw7NuIGRpYWwgY2lyY3VsYXIgYW5hbMOzZ2ljbyIsICJHZW7DqXJpY28iLCAiIl0sIFsiTWFuaWZvbGQgbmV1bcOhdGljbyA1LzIgU01DIGNvbiBzb2xlbm9pZGVzIiwgIk5ldW3DoXRpY28iLCAiQWx0YSIsICJBbG1hY8OpbiIsIDEsIDIsICI1LTcgZMOtYXMiLCAiU01DIiwgIkVzdGFudGUtRWxlY3Ryb3ZhbHZ1bGFzIiwgMCwgIk1hbmlmb2xkIDUgcG9zaWNpb25lcy8yIHbDrWFzIGNvbiBib2JpbmFzIDIyMFZBQyIsICJTTUMiLCAiIl0sIFsiQ29uZWN0b3JlcyBwdXNoLXRvLWNvbm5lY3QgbmV1bcOhdGljb3Mgc3VydGlkbyIsICJOZXVtw6F0aWNvIiwgIk1lZGlhIiwgIkFsbWFjw6luIiwgMTAsIDEsICIzLTUgZMOtYXMiLCAiR2Vuw6lyaWNvIiwgIkVzdGFudGUtRWxlY3Ryb3ZhbHZ1bGFzIiwgMCwgIlN1cnRpZG8gY29kb3MsIFQsIHJlY3RvcyA2LTgtMTBtbSB5IGFjZXJvIiwgIkdlbsOpcmljbyIsICIiXSwgWyJWw6FsdnVsYXMgZGUgYm9sYSAxLzQgbGF0w7NuIG1pbmkiLCAiTmV1bcOhdGljbyIsICJNZWRpYSIsICJBbG1hY8OpbiIsIDUsIDEwLCAiMy01IGTDrWFzIiwgIkdlbsOpcmljbyIsICJFc3RhbnRlLUVsZWN0cm92YWx2dWxhcyIsIDAsICJWw6FsdnVsYSBkZSBib2xhIDEvNCBsYXTDs24gcGFyYSBuZXVtw6F0aWNhIiwgIkdlbsOpcmljbyIsICIiXSwgWyJCb2JpbmEgc29sZW5vaWRlIGNvbmVjdG9yIERJTiBncmlzIDIyMFZBQyIsICJOZXVtw6F0aWNvIiwgIk1lZGlhIiwgIkFsbWFjw6luIiwgMiwgMiwgIjMtNSBkw61hcyIsICJHZW7DqXJpY28iLCAiRXN0YW50ZS1FbGVjdHJvdmFsdnVsYXMiLCAwLCAiQm9iaW5hIHJlcHVlc3RvIHbDoWx2dWxhIHNvbGVub2lkZSBjb25lY3RvciBESU4gZ3JpcyIsICJHZW7DqXJpY28iLCAiIl0sIFsiVsOhbHZ1bGEgc29sZW5vaWRlIEhVWVUgMjRWREMgcGFyYSBtYW5pZm9sZCIsICJOZXVtw6F0aWNvIiwgIk1lZGlhIiwgIkFsbWFjw6luIiwgMSwgMSwgIjUtNyBkw61hcyIsICJIVVlVIiwgIkVzdGFudGUtRWxlY3Ryb3ZhbHZ1bGFzIiwgMCwgIjI0VkRDIG1vbnRhamUgZW4gbWFuaWZvbGQiLCAiSFVZVSIsICIiXSwgWyJDYWJsZSBWaWFrb24gTm8uMTQgcm9qbyAxMDBtIiwgIkVsw6ljdHJpY28iLCAiQWx0YSIsICJBbG1hY8OpbiIsIDEsIDEsICIzLTUgZMOtYXMiLCAiVmlha29uIiwgIkVzdGFudGUtQ2FibGVzIiwgMCwgIkNvbmR1Y3RvciBlbMOpY3RyaWNvIDE0QVdHIHJvam8gOTDCsEMgMTAwbSIsICJWaWFrb24iLCAiTm8uMTQiXSwgWyJDaW50YSBhaXNsYW50ZSBlbMOpY3RyaWNhIDNNIiwgIkNvbnN1bWlibGUiLCAiQmFqYSIsICJBbG1hY8OpbiIsIDIsIDIsICIxLTMgZMOtYXMiLCAiM00iLCAiRXN0YW50ZS1Db25zdW1pYmxlcyIsIDAsICJDaW50YSBhaXNsYW50ZSBuZWdyYSB5IGJsYW5jYSAzTSIsICIzTSIsICIiXSwgWyJDaW50YSBmaWJyYSBkZSB2aWRyaW8gZWzDqWN0cmljYSIsICJDb25zdW1pYmxlIiwgIk1lZGlhIiwgIkFsbWFjw6luIiwgMSwgMSwgIjMtNSBkw61hcyIsICJHZW7DqXJpY28iLCAiRXN0YW50ZS1Db25zdW1pYmxlcyIsIDAsICJDaW50YSBmaWJyYSB2aWRyaW8gcGFyYSByZXNpc3RlbmNpYXMvYWx0YSB0ZW1wZXJhdHVyYSIsICJHZW7DqXJpY28iLCAiIl0sIFsiUm9kYW1pZW50byBkZSBib2xhcyBaU0cgc2VsbGFkbyIsICJNZWPDoW5pY28iLCAiTWVkaWEiLCAiQWxtYWPDqW4iLCAyLCA0LCAiMy01IGTDrWFzIiwgIlpTRyIsICJFc3RhbnRlLVJvZGFtaWVudG9zIiwgMCwgIlJvZGFtaWVudG8gYm9sYXMgWlNHIGNhamEgYXp1bCBzZWxsYWRvIiwgIlpTRyIsICIiXSwgWyJCbG9xdWUgY29udGFjdG8gWkIyLUJFMTAxIE5PIDEwQSIsICJFbMOpY3RyaWNvIiwgIk1lZGlhIiwgIkFsbWFjw6luIiwgMywgMywgIjMtNSBkw61hcyIsICJHZW7DqXJpY28iLCAiRXN0YW50ZS1Cb3RvbmVzIiwgMCwgIkJsb3F1ZSBjb250YWN0byBub3JtYWxtZW50ZSBhYmllcnRvIDIybW0iLCAiU2NobmVpZGVyL0NvbXAiLCAiWkIyLUJFMTAxIl0sIFsiQ29uZGVuc2Fkb3IgZGUgYXJyYW5xdWUgbW90b3IiLCAiRWzDqWN0cmljbyIsICJNZWRpYSIsICJBbG1hY8OpbiIsIDEsIDEsICI1LTcgZMOtYXMiLCAiR2Vuw6lyaWNvIiwgIkVzdGFudGUtVmFyaW9zIiwgMCwgIkNvbmRlbnNhZG9yIGVsZWN0cm9sw610aWNvIGNpbMOtbmRyaWNvIHBhcmEgYXJyYW5xdWUgZGUgbW90b3IiLCAiR2Vuw6lyaWNvIiwgIiJdLCBbIlRvcm5pbGxlcsOtYSBhdXRvcnJvc2NhbnRlIHN1cnRpZGEgKGJvbHNhKSIsICJDb25zdW1pYmxlIiwgIkJhamEiLCAiQWxtYWPDqW4iLCAyMCwgMjAwLCAiMS0yIGTDrWFzIiwgIkdlbsOpcmljbyIsICJFc3RhbnRlLUNvbnN1bWlibGVzIiwgMCwgIlRvcm5pbGxvcyBhdXRvLXBlcmZvcmFudGVzIGhleGFnb25hbGVzIHZhcmlvcyB0YW1hw7FvcyBlbiBib2xzYXMiLCAiR2Vuw6lyaWNvIiwgIiJdLCBbIlBhc3RhIHTDqXJtaWNhIEFyY3RpYyBNWC00IiwgIkNvbnN1bWlibGUiLCAiQmFqYSIsICJBbG1hY8OpbiIsIDEsIDIsICIzLTUgZMOtYXMiLCAiQXJjdGljIiwgIkVzdGFudGUtQ29uc3VtaWJsZXMiLCAwLCAiQ29tcHVlc3RvIHTDqXJtaWNvIGFsdGEgY29uZHVjdGl2aWRhZCB0dWJvIDRnIiwgIkFyY3RpYyIsICJNWC00Il0sIFsiQmxvcXVlcyB0ZXJtaW5hbGVzIFBMQyB2ZXJkZXMgY2FycmlsIERJTiIsICJFbMOpY3RyaWNvIiwgIk1lZGlhIiwgIkFsbWFjw6luIiwgMSwgMSwgIjMtNSBkw61hcyIsICJHZW7DqXJpY28iLCAiRXN0YW50ZS1UZXJtaW5hbGVzIiwgMCwgIkJsb3F1ZXMgdGVybWluYWxlcyB0aXBvIGNsYW1wIHZlcmRlcyBwYXJhIGNhcnJpbCBESU4iLCAiR2Vuw6lyaWNvIiwgIiJdXQ==").decode("utf-8"))
    for _r in _seed:
        try:
            con.execute("INSERT OR IGNORE INTO refacciones (nombre,categoria,criticidad,seccion,cant_min,stock_actual,tiempo_entrega,proveedor,ubicacion,costo,notas,marca,modelo) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", _r)
        except: pass
    con.commit()  # seed limpio 83 items
    # -- Deduplicar refacciones: fusionar duplicados por nombre sumando stock --
    try:
        dups = con.execute("""
            SELECT nombre, COUNT(*) as cnt, SUM(stock_actual) as total_stock,
                   MAX(cant_min) as max_min, MIN(id) as keep_id
            FROM refacciones GROUP BY nombre HAVING COUNT(*) > 1
        """).fetchall()
        for d in dups:
            nombre, cnt, total_stock, max_min, keep_id = d
            con.execute("UPDATE refacciones SET stock_actual=?, cant_min=? WHERE id=?",
                        (total_stock, max_min, keep_id))
            con.execute("DELETE FROM refacciones WHERE nombre=? AND id!=?", (nombre, keep_id))
        if dups:
            con.commit()  # dedup
    except: pass
    con.commit(); con.close()

def init_users_db():
    con = get_db()
    con.execute(
        """CREATE TABLE IF NOT EXISTS users
        (id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL,
         email TEXT NOT NULL UNIQUE, pin_hash TEXT NOT NULL,
         created_at TEXT DEFAULT (datetime('now')))"""
    )
    con.commit(); con.close()

init_db()
init_users_db()

# Seed usuario admin por defecto
try:
    import hashlib as _hl
    _con = get_db()
    _con.execute("INSERT OR IGNORE INTO users (nombre, email, pin_hash) VALUES (?,?,?)",
                 ("Mariano Leyva", "marianoleyva2608@gmail.com", _hl.sha256("260881".encode()).hexdigest()))
    _con.commit(); _con.close()
except: pass

# Migración: agregar imagen_url si no existe
try:
    _mcon = get_db()
    _mcon.execute("ALTER TABLE refacciones ADD COLUMN imagen_url TEXT DEFAULT ''")
    _mcon.commit(); _mcon.close()
except: pass

# Migración: agregar numero_parte (para QR) si no existe.
# Las refacciones ya existentes quedan con numero_parte='' (no se tocan/regeneran).
try:
    _npcon = get_db()
    _npcon.execute("ALTER TABLE refacciones ADD COLUMN numero_parte TEXT DEFAULT ''")
    _npcon.commit(); _npcon.close()
except: pass

# Seed imágenes de productos (UPDATE por nombre LIKE)
try:
    _icon = get_db()
    _img_seeds = [
        ('%TCN4S-24R%',   'https://ce8dc832c.cloudimg.io/v7/_cdn_/6A/12/90/00/0/598438_1.jpg'),
        ('%AT8N%',         'https://ce8dc832c.cloudimg.io/v7/_cdn_/08/E1/90/00/0/597632_1.jpg'),
        ('%NXC-32%',       'https://www.sparegenie.com/cdn/shop/products/10001699.jpg?v=1657890170'),
        ('%NR2-25%',       'https://media.rs-online.com/image/upload/b_auto,c_pad,dpr_1,f_auto,h_399,q_auto,w_399/R1948771-01'),
        ('%SC-4-1%',       'https://www.yuengkao.com/images/products/1385961666_mc_sc-4-1_2.jpg'),
        ('%60.12.8.230%',  'https://ce8dc832c.cloudimg.io/v7/_cdn_/C8/2C/50/00/0/377484_1.jpg'),
        ('%ZB2-BE101%',    'https://ce8dc832c.cloudimg.io/v7/_cdn_/6D/5F/50/00/0/390614_1.jpg'),
    ]
    for _pat, _url in _img_seeds:
        _icon.execute("UPDATE refacciones SET imagen_url=? WHERE nombre LIKE ? AND (imagen_url IS NULL OR imagen_url='')", (_url, _pat))
    _icon.commit(); _icon.close()
except: pass

def hash_pin(pin):
    return hashlib.sha256(pin.encode()).hexdigest()

@app.route('/api/users/register', methods=['POST'])
def register_user():
    d = request.json
    nombre = d.get('nombre','').strip()
    email  = d.get('email','').strip().lower()
    pin    = d.get('pin','').strip()
    if not nombre or not email or not pin or len(pin) < 4:
        return jsonify({'error': 'Nombre, email y PIN (minimo 4 caracteres) requeridos'}), 400
    try:
        con = get_db()
        con.execute('INSERT INTO users (nombre, email, pin_hash) VALUES (?,?,?)',
                    (nombre, email, hash_pin(pin)))
        con.commit(); con.close()
        return jsonify({'ok': True, 'nombre': nombre, 'email': email})
    except Exception as e:
        if 'UNIQUE' in str(e):
            return jsonify({'error': 'Este email ya esta registrado'}), 409
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/login', methods=['POST'])
def login_user():
    d = request.json
    email = d.get('email','').strip().lower()
    pin   = d.get('pin','').strip()
    con = get_db()
    row = con.execute('SELECT id, nombre, email FROM users WHERE email=? AND pin_hash=?',
                      (email, hash_pin(pin))).fetchone()
    con.close()
    if not row:
        return jsonify({'error': 'Email o PIN incorrecto'}), 401
    return jsonify({'ok': True, 'id': row['id'], 'nombre': row['nombre'], 'email': row['email']})

@app.route('/api/users/verify', methods=['POST'])
def verify_user():
    d = request.json
    email = d.get('email','').strip().lower()
    pin   = d.get('pin','').strip()
    con = get_db()
    row = con.execute('SELECT id, nombre, email FROM users WHERE email=? AND pin_hash=?',
                      (email, hash_pin(pin))).fetchone()
    con.close()
    if not row:
        return jsonify({'ok': False, 'error': 'PIN incorrecto'}), 200
    return jsonify({'ok': True, 'nombre': row['nombre'], 'email': row['email']})

@app.route('/api/users', methods=['GET'])
def list_users():
    con = get_db()
    rows = con.execute('SELECT id, nombre, email, created_at FROM users ORDER BY nombre').fetchall()
    con.close()
    return jsonify([dict(r) for r in rows])

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.page import PageMargins
except ImportError:
    os.system("pip install openpyxl -q")
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.page import PageMargins

MONTH_COLS = {"ENE":"B","FEB":"C","MAR":"D","ABR":"E","MAY":"F","JUN":"G",
              "JUL":"H","AGO":"I","SEP":"J","OCT":"K","NOV":"L","DIC":"M"}

def _wc(ws, row, col, value, alignment=None):
    cell = ws.cell(row=row, column=col)
    try:
        cell.value = value
    except AttributeError:
        return
    if alignment:
        cell.alignment = alignment

def apply_checklist_data(ws, data):
    SHEET_SIG = {
        'TF 1':45,'TF 2':45,'TF 3':45,'TF 4':45,'TF 5':45,'TF 6':45,'TF 7':45,
        'prensa 1':40,'Prensa 2':40,
        'Suajadora 1':26,'Suajadora 2':26,'Suajadora 5':26,'TCU':26,'Chiller':27,
        'Laminator':37,'Slitter Rewinder':37,'Komatsu OBS45':35,
        'Rotary Press':37,'Prensa PL':32,'IMESA':31,'Single Knife':31,
        'Hojeadora Robust':31,'Gapcutter':30,'Calender K1 Easy':41,
    }
    sheet_name = data.get('sheet', '')
    sig_row = SHEET_SIG.get(sheet_name) or data.get('sig_row')
    cal_row = data.get('cal_data_row')
    max_item_row = (sig_row - 1) if sig_row else 9999

    for item in data.get('items', []):
        row, st, cm = item['row'], item.get('status',''), item.get('comment','')
        if row > max_item_row:
            continue
        _wc(ws, row, 11, '( v )' if st == 'ok' else '(   )')
        _wc(ws, row, 12, '( v )' if st == 'ng' else '(   )')
        if cm:
            _wc(ws, row, 13, cm)

    months = set(data.get('month', []))
    if cal_row:
        for m, col in MONTH_COLS.items():
            _wc(ws, cal_row, ord(col)-ord('A')+1, '( v )' if m in months else '(   )')
        v = data.get('voltaje', {})
        if sheet_name == 'Rotary Press':
            if v.get('l1'):  _wc(ws, cal_row, 14, v['l1'])
            if v.get('vac'): _wc(ws, cal_row, 15, v['vac'])
        else:
            for col_n, key in [(14,'l1'),(15,'l2'),(16,'l3'),(17,'vac')]:
                if v.get(key):
                    _wc(ws, cal_row, col_n, v[key])

    if sig_row:
        tec = data.get('tecnico','_______________')
        fec = data.get('fecha','_______________')
        sup = data.get('supervisor','_______________')
        firma_row = sig_row + 1
        _wc(ws, sig_row, 2,  "Realizo:")
        _wc(ws, sig_row, 8,  "Fecha: " + fec)
        fecha_cell = ws.cell(row=sig_row, column=8)
        fecha_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        fecha_cell.font = Font(bold=True, color='FFFFFF', size=11)
        fecha_cell.fill = PatternFill(fill_type='solid', fgColor='FF1A5C2A')
        _wc(ws, sig_row, 11, "Supervisor de Produccion:")

        tec_sig = data.get('sigTecnico') or {}
        sup_sig = data.get('sigSupervisor') or {}

        if isinstance(tec_sig, dict) and not tec_sig.get('nombre') and tec and tec != '_______________':
            tec_sig = dict(tec_sig); tec_sig['nombre'] = tec
        if isinstance(sup_sig, dict) and not sup_sig.get('nombre') and sup and sup != '_______________':
            sup_sig = dict(sup_sig); sup_sig['nombre'] = sup

        has_tec = isinstance(tec_sig, dict) and bool(tec_sig.get('nombre')) and bool(tec_sig.get('fecha'))
        has_sup = isinstance(sup_sig, dict) and bool(sup_sig.get('nombre')) and bool(sup_sig.get('fecha'))

        if has_tec or has_sup:
            ws.row_dimensions[firma_row].height = 70
            ws.row_dimensions[firma_row + 1].height = 0  # ocultar fila extra

        def write_pki(col, sig):
            pki_text = (sig.get('nombre','') + "\n"
                "Fecha:    " + sig.get('fecha','') + "\n"
                "Serie:    " + sig.get('serie','') + "\n"
                "Emisor:   AD-PACK Mexico")
            cell = ws.cell(row=firma_row, column=col)
            try:
                cell.value = pki_text
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
                cell.font = Font(name='Calibri', size=10, bold=True, color='FF000000')
            except AttributeError:
                pass

        if has_tec:
            write_pki(2, tec_sig)
        if has_sup:
            write_pki(11, sup_sig)

        # --- UNIFICAR PAGE SETUP en todos los formatos ---
        last_print_row = firma_row  # no incluir fila extra
        start_print_row = 2 if sheet_name.startswith('TF') else 1
        ws.print_area = '$B${}:$Q${}'.format(start_print_row, last_print_row)
        ws.page_setup.orientation = 'portrait'
        if sheet_name == 'Rotary Press':
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 1
            ws.page_setup.fitToWidth = 1
        else:
            ws.page_setup.scale = 44
        ws.page_margins = PageMargins(
            left=0.70, right=0.70, top=0.75, bottom=0.75,
            header=0.3, footer=0.3
        )

@app.route('/api/reports', methods=['GET'])
def api_get_reports():
    con = get_db()
    rows = con.execute('SELECT data FROM reports ORDER BY id DESC').fetchall()
    con.close()
    return jsonify([json.loads(r['data']) for r in rows])

@app.route('/api/reports', methods=['POST'])
def api_save_report():
    r = request.json
    con = get_db()
    con.execute('DELETE FROM reports WHERE machine_id=? AND fecha=?', (r.get('machineId',''), r.get('fecha','')))
    con.execute('INSERT INTO reports (id, machine_id, fecha, data) VALUES (?,?,?,?)',
                (r['id'], r.get('machineId',''), r.get('fecha',''), json.dumps(r)))
    con.commit(); con.close()
    return jsonify({'ok': True})

@app.route('/api/reports/<int:report_id>', methods=['DELETE'])
def api_delete_report(report_id):
    con = get_db()
    con.execute('DELETE FROM reports WHERE id=?', (report_id,))
    con.commit(); con.close()
    return jsonify({'ok': True})
@app.route('/api/reports/export', methods=['GET'])
def api_export_reports():
    import datetime
    con = get_db()
    rows = con.execute('SELECT data FROM reports ORDER BY id DESC').fetchall()
    con.close()
    data = [json.loads(r['data']) for r in rows]
    fname = 'reportes_backup_' + datetime.datetime.now().strftime('%Y%m%d_%H%M%S') + '.json'
    resp = app.response_class(
        response=json.dumps(data, ensure_ascii=False, indent=2),
        status=200, mimetype='application/json'
    )
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp

@app.route('/api/reports/import', methods=['POST'])
def api_import_reports():
    reports = request.json
    if not isinstance(reports, list):
        return jsonify({'error': 'Se esperaba una lista'}), 400
    con = get_db()
    imported = 0
    for r in reports:
        try:
            con.execute('INSERT OR IGNORE INTO reports (id, machine_id, fecha, data) VALUES (?,?,?,?)',
                        (r['id'], r.get('machineId',''), r.get('fecha',''), json.dumps(r)))
            imported += 1
        except Exception:
            pass
    con.commit(); con.close()
    return jsonify({'ok': True, 'imported': imported})


@app.route('/')
def index():
    html = open('index.html', encoding='utf-8').read()
    return html, 200, {'Content-Type':'text/html; charset=utf-8', 'Cache-Control':'no-store, no-cache, must-revalidate', 'Pragma':'no-cache'}

@app.route('/generar', methods=['POST'])
def generar():
    try:
        data = request.json
        file_key = data.get('file', 'termoformado')
        template = 'MTTO_Preventivo_Termoformado_2026.xlsx' if file_key == 'termoformado' else 'MTTO_Preventivo_Conversion_2026.xlsx'
        wb = openpyxl.load_workbook(template)
        ws = wb[data['sheet']]
        apply_checklist_data(ws, data)
        for s in [n for n in wb.sheetnames if n != ws.title]:
            del wb[s]
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = "REPORTE_" + data.get('machineId','EQ') + "_" + data.get('fecha','').replace('/','-') + ".xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        import traceback; traceback.print_exc()
        return {'error': str(e)}, 500

@app.route('/generar_pdf', methods=['POST'])
def generar_pdf():
    import tempfile, subprocess, shutil
    tmp_dir = tempfile.mkdtemp()
    try:
        data = request.json
        file_key = data.get('file', 'termoformado')
        template = 'MTTO_Preventivo_Termoformado_2026.xlsx' if file_key == 'termoformado' else 'MTTO_Preventivo_Conversion_2026.xlsx'
        wb = openpyxl.load_workbook(template)
        ws = wb[data['sheet']]
        apply_checklist_data(ws, data)
        for s in [s for s in wb.sheetnames if s != ws.title]:
            del wb[s]
        xlsx_path = os.path.join(tmp_dir, 'reporte.xlsx')
        wb.save(xlsx_path)
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'pdf', '--outdir', tmp_dir, xlsx_path],
            capture_output=True, text=True, timeout=60)
        if result.returncode != 0:
            raise RuntimeError("LibreOffice error: " + result.stderr)
        pdf_path = os.path.join(tmp_dir, 'reporte.pdf')
        if not os.path.exists(pdf_path):
            raise RuntimeError("PDF no generado por LibreOffice")
        with open(pdf_path, 'rb') as f:
            pdf_bytes = f.read()
        fname = "REPORTE_" + data.get('machineId','EQ') + "_" + data.get('fecha','').replace('/','-') + ".pdf"
        return send_file(io.BytesIO(pdf_bytes), as_attachment=True, download_name=fname, mimetype='application/pdf')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@app.route('/api/work-orders', methods=['GET'])
def api_get_work_orders():
    con = get_db()
    rows = con.execute('SELECT * FROM work_orders ORDER BY id DESC').fetchall()
    con.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/work-orders', methods=['POST'])
def api_save_work_order():
    data = request.json
    con = get_db()
    last = con.execute('SELECT MAX(CAST(numero AS INTEGER)) as mx FROM work_orders').fetchone()
    next_num = (last['mx'] or 257) + 1
    numero = str(next_num).zfill(4)
    con.execute(
        'INSERT INTO work_orders (numero,solicitante,fecha,equipo,planta,tipo,estatus,hora_inicio,hora_termino,'
        'tiempo_paro,descripcion_falla,actividad_realizada,refaccion,observaciones,'
        'firma_solicitante,firma_recibe,firma_liberacion,fotos) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
        (numero, data.get('solicitante',''), data.get('fecha',''), data.get('equipo',''),
         data.get('planta',''), data.get('tipo',''), data.get('estatus',''),
         data.get('hora_inicio',''), data.get('hora_termino',''), data.get('tiempo_paro',''),
         data.get('descripcion_falla',''), data.get('actividad_realizada',''),
         data.get('refaccion',''), data.get('observaciones',''),
         data.get('firma_solicitante',''), data.get('firma_recibe',''), data.get('firma_liberacion',''), data.get('fotos','[]'))
    )
    con.commit()
    new_id = con.execute('SELECT last_insert_rowid()').fetchone()[0]
    con.close()
    return jsonify({'ok': True, 'id': new_id, 'numero': numero})

@app.route('/api/work-orders/<int:order_id>', methods=['DELETE'])
def api_delete_work_order(order_id):
    con = get_db()
    con.execute('DELETE FROM work_orders WHERE id=?', (order_id,))
    con.commit(); con.close()
    return jsonify({'ok': True})

@app.route('/api/work-orders/next-number', methods=['GET'])
def api_next_order_number():
    con = get_db()
    last = con.execute('SELECT MAX(CAST(numero AS INTEGER)) as mx FROM work_orders').fetchone()
    next_num = (last['mx'] or 257) + 1
    con.close()
    return jsonify({'numero': str(next_num).zfill(4)})

@app.route('/version')
def version():
    return jsonify({'commit': 'a1f9c33', 'fix': 'unified_page_setup_all_formats'})


# ── PDF ORDEN DE TRABAJO ──────────────────────────────────────────
@app.route('/api/work-orders/<int:order_id>/pdf')
def api_orden_pdf(order_id):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfgen import canvas as pdfcanvas
    import io, json, base64

    con = get_db()
    o = con.execute('SELECT * FROM work_orders WHERE id=?',(order_id,)).fetchone()
    con.close()
    if not o: return jsonify({'error':'not found'}),404
    o = dict(o)

    buf = io.BytesIO()
    W, H = letter  # 612 x 792 pts
    c = pdfcanvas.Canvas(buf, pagesize=letter)

    GREEN = colors.HexColor('#1b5e20')
    RED   = colors.HexColor('#c62828')
    LGRAY = colors.HexColor('#f5f5f5')
    BLACK = colors.black

    margin = 1.5*cm
    w = W - 2*margin
    y = H - margin

    def line(y): c.setStrokeColor(colors.HexColor('#cccccc')); c.line(margin, y, margin+w, y)
    def chk(val, opt): return '[v]' if val==opt else '[ ]'
    def txt(t,x,yy,size=9,bold=False,color=BLACK):
        c.setFont('Helvetica-Bold' if bold else 'Helvetica', size)
        c.setFillColor(color)
        c.drawString(x,yy,str(t or ''))

    # ── BORDER ──
    c.setStrokeColor(GREEN); c.setLineWidth(2)
    c.rect(margin-2, margin-2, w+4, H-2*margin+4)

    # ── HEADER ──
    y -= 0.3*cm
    # AD-PACK box
    c.setStrokeColor(GREEN); c.setLineWidth(1.5)
    c.rect(margin+2, y-1.2*cm, 2.8*cm, 1.4*cm)
    c.setFont('Helvetica-Bold',14); c.setFillColor(GREEN)
    c.drawString(margin+8, y-0.5*cm, 'ad')
    c.setFont('Helvetica-Bold',9)
    c.drawString(margin+6, y-0.9*cm, 'AD-PACK')

    # Title
    c.setFont('Helvetica-Bold',14); c.setFillColor(GREEN)
    c.drawCentredString(W/2, y-0.3*cm, 'ORDEN DE TRABAJO PARA MANTENIMIENTO')

    # OT Number
    c.setFont('Helvetica-Bold',22); c.setFillColor(RED)
    c.drawRightString(margin+w-2, y-0.5*cm, str(o['numero']))

    y -= 1.6*cm
    c.setStrokeColor(GREEN); c.setLineWidth(1.5)
    c.line(margin, y, margin+w, y)
    y -= 0.1*cm

    # ── ROW 1: Solicitante / Fecha / Equipo ──
    y -= 0.55*cm
    txt('SOLICITANTE:', margin+2, y, 8, True)
    txt(o['solicitante'] or '', margin+2.5*cm, y, 9)
    txt('FECHA:', margin+w*0.55, y, 8, True)
    txt(o['fecha'] or '', margin+w*0.55+1.5*cm, y, 9)
    txt('EQUIPO:', margin+w*0.75, y, 8, True)
    txt(o['equipo'] or '', margin+w*0.75+1.5*cm, y, 9)
    c.setStrokeColor(LGRAY); c.line(margin, y-3, margin+w, y-3)

    # ── ROW 2: Planta / Tipo / Estatus ──
    y -= 0.7*cm
    txt('PLANTA:', margin+2, y, 8, True)
    txt(chk(o['planta'],'TERMOFORMADO')+' TERMOFORMADO', margin+1.8*cm, y, 9)
    txt(chk(o['planta'],'CONVERSIÓN')+' CONVERSIÓN', margin+5.5*cm, y, 9)
    txt('TIPO:', margin+w*0.48, y, 8, True)
    txt(chk(o['tipo'],'CORRECTIVO')+' CORRECTIVO', margin+w*0.48+1*cm, y, 9)
    txt(chk(o['tipo'],'PREVENTIVO')+' PREVENTIVO', margin+w*0.67, y, 9)
    txt('ESTATUS:', margin+w*0.83, y, 8, True)
    txt(chk(o['estatus'],'CERRADA')+' CERRADA', margin+w*0.83+1.5*cm, y, 9)
    y -= 0.4*cm
    txt(chk(o['estatus'],'ABIERTA')+' ABIERTA', margin+w*0.83+1.5*cm, y, 9)
    c.setStrokeColor(LGRAY); c.line(margin, y-3, margin+w, y-3)

    # ── ROW 3: Horas ──
    y -= 0.7*cm
    txt('HORA DE INICIO:', margin+2, y, 8, True)
    txt(o['hora_inicio'] or '_____', margin+3*cm, y, 10, True)
    txt('HORA DE TÉRMINO:', margin+w*0.35, y, 8, True)
    txt(o['hora_termino'] or '_____', margin+w*0.35+3.5*cm, y, 10, True)
    txt('TIEMPO DE PARO:', margin+w*0.67, y, 8, True)
    txt(o['tiempo_paro'] or '_____', margin+w*0.67+3*cm, y, 10, True, RED)
    c.setStrokeColor(GREEN); c.setLineWidth(1)
    y -= 0.3*cm; c.line(margin, y, margin+w, y)

    def text_box(label, value, yy, height=2.2*cm):
        c.setStrokeColor(BLACK); c.setLineWidth(0.5)
        c.rect(margin, yy-height, w, height)
        txt(label+':', margin+4, yy-0.35*cm, 8, True)
        if value:
            # wrap text
            from reportlab.pdfbase.pdfmetrics import stringWidth
            words = str(value).split()
            line_txt = ''; line_y = yy-0.7*cm; max_w = w-10
            for word in words:
                test = line_txt+(' ' if line_txt else '')+word
                if stringWidth(test,'Helvetica',9) < max_w:
                    line_txt = test
                else:
                    if line_txt: c.setFont('Helvetica',9); c.setFillColor(BLACK); c.drawString(margin+4, line_y, line_txt)
                    line_txt = word; line_y -= 0.45*cm
            if line_txt: c.setFont('Helvetica',9); c.setFillColor(BLACK); c.drawString(margin+4, line_y, line_txt)
        return yy - height - 0.2*cm

    y -= 0.1*cm
    y = text_box('DESCRIPCIÓN DE LA FALLA', o.get('descripcion_falla',''), y, 2.5*cm)
    y = text_box('ACTIVIDAD REALIZADA', o.get('actividad_realizada',''), y, 2.5*cm)
    y = text_box('REFACCIÓN UTILIZADA', o.get('refaccion',''), y, 1.8*cm)
    y = text_box('OBSERVACIONES', o.get('observaciones',''), y, 1.6*cm)

    # ── FOTOS ──
    try:
        imgs = json.loads(o.get('fotos','[]'))
        if imgs:
            y -= 0.2*cm
            txt('EVIDENCIA FOTOGRÁFICA ('+str(len(imgs))+')', margin+2, y, 8, True, GREEN)
            y -= 0.2*cm
            ix = margin
            for b64 in imgs[:5]:
                if ',' in b64: b64 = b64.split(',')[1]
                img_buf = io.BytesIO(base64.b64decode(b64))
                img = Image(img_buf, width=2.5*cm, height=2.5*cm)
                img.drawOn(c, ix, y-2.7*cm)
                ix += 2.7*cm
            y -= 3*cm
    except: pass

    # ── FIRMAS ──
    y -= 0.5*cm
    sig_w = w/3
    def draw_sig(label, data, sx):
        sd = None
        try:
            if data: sd = json.loads(data)
        except: pass
        c.setStrokeColor(BLACK); c.setLineWidth(0.8)
        c.line(sx+0.5*cm, y, sx+sig_w-0.5*cm, y)
        if sd and sd.get('nombre'):
            txt(sd['nombre'], sx+0.5*cm, y+0.15*cm, 7, True, GREEN)
            txt('✓ FIRMA ELECTRÓNICA VÁLIDA', sx+0.5*cm, y-0.35*cm, 6, False, GREEN)
            txt('CERT: '+str(sd.get('certCode','')), sx+0.5*cm, y-0.65*cm, 6)
            txt(str(sd.get('timestamp','')), sx+0.5*cm, y-0.9*cm, 6)
        c.setFont('Helvetica-Bold',8); c.setFillColor(BLACK)
        c.drawCentredString(sx+sig_w/2, y-1.3*cm, label)

    draw_sig('SOLICITANTE', o.get('firma_solicitante'), margin)
    draw_sig('RECIBE ORDEN DE TRABAJO', o.get('firma_recibe'), margin+sig_w)
    draw_sig('LIBERACIÓN DE PRODUCCIÓN', o.get('firma_liberacion'), margin+2*sig_w)

    # Footer
    c.setFont('Helvetica',7); c.setFillColor(colors.HexColor('#999999'))
    c.drawRightString(margin+w, margin+0.3*cm, 'A9.F4 Rev. 01')

    c.save()
    buf.seek(0)
    from flask import send_file
    return send_file(buf, mimetype='application/pdf',
                     download_name=f'OT-{o["numero"]}.pdf',
                     as_attachment=False)

# ══════════════════════════════════════════════════════════
#  REQUISICIONES DE COMPRA
# ══════════════════════════════════════════════════════════

def init_req_db():
    con = get_db()
    con.execute('''CREATE TABLE IF NOT EXISTS requisiciones
        (id TEXT PRIMARY KEY,
         folio INTEGER,
         fecha TEXT,
         solicitante TEXT,
         planta TEXT,
         departamento TEXT,
         tipo TEXT,
         data TEXT,
         created_at TEXT DEFAULT (datetime('now')))''')
    con.commit(); con.close()

init_req_db()

@app.route('/api/requisiciones', methods=['GET'])
def get_requisiciones():
    try:
        init_req_db()
        con = get_db()
        rows = con.execute('SELECT data FROM requisiciones ORDER BY folio DESC').fetchall()
        con.close()
        return jsonify([json.loads(r['data']) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/requisiciones', methods=['POST'])
def save_requisicion():
    d = request.json
    con = get_db()
    if d.get('id'):
        existing = con.execute('SELECT folio FROM requisiciones WHERE id=?', (d['id'],)).fetchone()
        folio = existing['folio'] if existing else None
    else:
        folio = None
    if not folio:
        last = con.execute('SELECT MAX(folio) as m FROM requisiciones').fetchone()
        folio = (last['m'] or 0) + 1
        d['id'] = d.get('id') or (str(folio).zfill(4))
    d['folio'] = folio
    con.execute('''INSERT OR REPLACE INTO requisiciones
        (id, folio, fecha, solicitante, planta, departamento, tipo, data)
        VALUES (?,?,?,?,?,?,?,?)''',
        (d['id'], folio, d.get('fecha',''), d.get('solicitante',''),
         d.get('planta',''), d.get('departamento',''), d.get('tipo','Normal'),
         json.dumps(d, ensure_ascii=False)))
    con.commit(); con.close()
    return jsonify({'ok': True, 'id': d['id'], 'folio': folio})

@app.route('/api/requisicion/<rid>/firmar', methods=['POST'])
def firmar_requisicion(rid):
    d = request.json
    key = d.get('key')   # 'realizo', 'reviso', 'aprobo'
    firma = d.get('firma')
    if key not in ('realizo','reviso','aprobo') or not firma:
        return jsonify({'ok': False, 'error': 'Datos inválidos'}), 400
    con = get_db()
    row = con.execute('SELECT data FROM requisiciones WHERE id=?', (rid,)).fetchone()
    if not row:
        con.close()
        return jsonify({'ok': False, 'error': 'No encontrado'}), 404
    o = json.loads(row['data'])
    if 'firmas' not in o or not isinstance(o['firmas'], dict):
        o['firmas'] = {}
    o['firmas'][key] = firma
    con.execute('UPDATE requisiciones SET data=? WHERE id=?', (json.dumps(o, ensure_ascii=False), rid))
    con.commit(); con.close()
    return jsonify({'ok': True})

@app.route('/api/requisicion/<rid>/pdf')
def requisicion_pdf(rid):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfgen import canvas as pdfcanvas

    con = get_db()
    row = con.execute('SELECT data FROM requisiciones WHERE id=?', (rid,)).fetchone()
    con.close()
    if not row:
        return jsonify({'error': 'No encontrado'}), 404
    o = json.loads(row['data'])

    buf = io.BytesIO()
    W, H = letter
    c = pdfcanvas.Canvas(buf, pagesize=letter)
    margin = 1.5*cm
    w = W - 2*margin
    GREEN = colors.HexColor('#1a5c2a')
    BLACK = colors.black
    GRAY  = colors.HexColor('#888888')
    LGRAY = colors.HexColor('#cccccc')

    def txt(t, x, yy, size=8, bold=False, color=BLACK, align='left'):
        c.setFont('Helvetica-Bold' if bold else 'Helvetica', size)
        c.setFillColor(color)
        if align == 'center':
            c.drawCentredString(x, yy, str(t or ''))
        elif align == 'right':
            c.drawRightString(x, yy, str(t or ''))
        else:
            c.drawString(x, yy, str(t or ''))

    y = H - margin

    # ── LOGO ──
    c.setStrokeColor(GREEN); c.setLineWidth(1.5)
    c.rect(margin, y-1.2*cm, 2.5*cm, 1.3*cm)
    c.setFont('Helvetica-Bold', 14); c.setFillColor(GREEN)
    c.drawString(margin+4, y-0.5*cm, 'ad')
    c.setFont('Helvetica-Bold', 8); c.setFillColor(GREEN)
    c.drawString(margin+4, y-0.9*cm, 'AD-PACK')

    # ── TITLE ──
    c.setFont('Helvetica-Bold', 13); c.setFillColor(BLACK)
    c.drawCentredString(W/2, y-0.35*cm, 'Requisición de compra')

    # ── DOC INFO top-right ──
    folio = str(o.get('folio', '')).zfill(4)
    bx = margin + w - 4.5*cm
    c.setStrokeColor(LGRAY); c.setLineWidth(0.5)
    c.rect(bx, y-1.2*cm, 4.5*cm, 1.3*cm)
    c.line(bx, y-0.4*cm, bx+4.5*cm, y-0.4*cm)
    c.line(bx+2.2*cm, y-1.2*cm, bx+2.2*cm, y)
    txt('No. de Documento', bx+4, y-0.15*cm, 6, True)
    txt('Revisión', bx+2.2*cm+4, y-0.15*cm, 6, True)
    txt('REQ-' + folio, bx+4, y-0.75*cm, 9, True, GREEN)
    txt(o.get('fecha', ''), bx+2.2*cm+4, y-0.75*cm, 7)

    y -= 1.4*cm

    # ── ROW 1: Fecha / Solicitante / Planta / Depto ──
    c.setStrokeColor(LGRAY); c.setLineWidth(0.5)
    row_h = 0.55*cm
    c.rect(margin, y-row_h, w, row_h)
    cols = [2.5*cm, w*0.35, w*0.55, w*0.75]
    for cx in cols:
        c.line(margin+cx, y, margin+cx, y-row_h)
    txt('FECHA:', margin+4, y-row_h+4, 7, True)
    txt(o.get('fecha',''), margin+1.2*cm, y-row_h+4, 7)
    txt('SOLICITANTE:', margin+cols[0]+4, y-row_h+4, 7, True)
    txt(o.get('solicitante',''), margin+cols[0]+2.3*cm, y-row_h+4, 7)
    txt('PLANTA:', margin+cols[1]+4, y-row_h+4, 7, True)
    txt(o.get('planta',''), margin+cols[1]+1.4*cm, y-row_h+4, 7)
    txt('DEPARTAMENTO:', margin+cols[2]+4, y-row_h+4, 7, True)
    txt(o.get('departamento',''), margin+cols[2]+2.5*cm, y-row_h+4, 7)

    y -= row_h

    # ── ROW 2: Tipo solicitud ──
    c.rect(margin, y-row_h, w, row_h)
    c.line(margin+w*0.55, y, margin+w*0.55, y-row_h)
    tipo = o.get('tipo', 'Normal')
    txt('TIPO DE SOLICITUD:', margin+4, y-row_h+4, 7, True)
    txt('[v]' if tipo=='Normal' else '[ ]', margin+3.8*cm, y-row_h+4, 7)
    txt('Normal', margin+4.4*cm, y-row_h+4, 7)
    txt('[v]' if tipo=='Urgente' else '[ ]', margin+5.5*cm, y-row_h+4, 7)
    txt('Urgente', margin+6.1*cm, y-row_h+4, 7, False, colors.HexColor('#c62828'))

    y -= row_h + 0.1*cm

    # ── TABLE HEADER ──
    col_widths = [0.55*cm, 1.4*cm, 1.5*cm, 4.5*cm, 3.0*cm, 3.2*cm, 3.3*cm]
    # adjust last col to fill width
    col_widths[-1] = w - sum(col_widths[:-1])
    col_x = [margin]
    for cw in col_widths[:-1]:
        col_x.append(col_x[-1] + cw)

    hdr_h = 1.0*cm
    c.setFillColor(colors.HexColor('#e8f5e9'))
    c.rect(margin, y-hdr_h, w, hdr_h, fill=1, stroke=1)
    c.setStrokeColor(LGRAY)
    for cx in col_x[1:]:
        c.line(cx, y, cx, y-hdr_h)

    headers = ['#','CANTIDAD','UNIDAD\n(pieza,caja,\nmetros,litros,\nrollos etc.)','DESCRIPCIÓN\n(nombre,modelo,marca,\nNo.catálogo,color,medidas)','APLICACIÓN\nO USO','PROVEEDOR SUGERIDO\n(nombre,contacto,\ntelefono/correo)','PROVEEDOR ALTERNO\n(nombre,contacto,\ntelefono/correo)']
    c.setFont('Helvetica-Bold', 6); c.setFillColor(BLACK)
    for i, (htext, cx, cw) in enumerate(zip(headers, col_x, col_widths)):
        lines = htext.split('\n')
        ly = y - 0.15*cm
        for ln in lines:
            c.drawCentredString(cx + cw/2, ly, ln)
            ly -= 0.22*cm

    y -= hdr_h

    # ── TABLE ROWS (10 items) ──
    items = o.get('items', [{}]*10)
    while len(items) < 10:
        items.append({})

    row_h_item = 0.7*cm
    for idx, item in enumerate(items[:10]):
        ry = y - row_h_item
        if idx % 2 == 0:
            c.setFillColor(colors.HexColor('#fafffe'))
        else:
            c.setFillColor(colors.white)
        c.rect(margin, ry, w, row_h_item, fill=1, stroke=0)
        c.setStrokeColor(LGRAY); c.setLineWidth(0.4)
        c.rect(margin, ry, w, row_h_item, fill=0, stroke=1)
        for cx in col_x[1:]:
            c.line(cx, y, cx, ry)

        c.setFont('Helvetica', 7); c.setFillColor(BLACK)
        cy = ry + row_h_item*0.35
        c.drawCentredString(col_x[0]+col_widths[0]/2, cy, str(idx+1))
        c.drawCentredString(col_x[1]+col_widths[1]/2, cy, str(item.get('cantidad','')))
        c.drawCentredString(col_x[2]+col_widths[2]/2, cy, str(item.get('unidad','')))
        c.drawString(col_x[3]+3, cy, str(item.get('descripcion',''))[:55])
        c.drawString(col_x[4]+3, cy, str(item.get('aplicacion',''))[:30])
        c.drawString(col_x[5]+3, cy, str(item.get('proveedor',''))[:28])
        c.drawString(col_x[6]+3, cy, str(item.get('proveedor_alt',''))[:28])
        y -= row_h_item

    y -= 0.2*cm

    # ── JUSTIFICACIÓN ──
    c.setStrokeColor(LGRAY); c.setLineWidth(0.5)
    just_h = 2.0*cm
    c.rect(margin, y-just_h, w, just_h)
    txt('JUSTIFICACIÓN (necesidad de la compra del material):', margin+4, y-0.25*cm, 7, True)
    just_txt = o.get('justificacion', '')
    c.setFont('Helvetica', 8); c.setFillColor(BLACK)
    # wrap text
    words = just_txt.split()
    line_txt = ''; ly = y - 0.55*cm; max_w = w - 10
    from reportlab.pdfbase.pdfmetrics import stringWidth
    for word in words:
        test = line_txt + (' ' if line_txt else '') + word
        if stringWidth(test, 'Helvetica', 8) < max_w:
            line_txt = test
        else:
            if line_txt: c.drawString(margin+4, ly, line_txt)
            line_txt = word; ly -= 0.35*cm
    if line_txt: c.drawString(margin+4, ly, line_txt)

    y -= just_h + 0.5*cm

    # ── FIRMAS ──
    sig_w = w / 3
    firmas = o.get('firmas', {})
    labels = ['Realizó (Nombre y firma)', 'Revisó (Nombre y firma)', 'Aprobó (Nombre y firma)']
    keys   = ['realizo', 'reviso', 'aprobo']
    for i, (label, key) in enumerate(zip(labels, keys)):
        sx = margin + i * sig_w
        c.setStrokeColor(BLACK); c.setLineWidth(0.8)
        c.line(sx+0.5*cm, y, sx+sig_w-0.5*cm, y)
        sig = firmas.get(key, {})
        if isinstance(sig, dict) and sig.get('nombre'):
            c.setFont('Helvetica-Bold', 7); c.setFillColor(GREEN)
            c.drawString(sx+0.5*cm, y+0.15*cm, sig.get('nombre',''))
            c.setFont('Helvetica', 6); c.setFillColor(GREEN)
            c.drawString(sx+0.5*cm, y-0.3*cm, '[v] FIRMA ELECTRONICA VALIDA')
        c.setFont('Helvetica-Bold', 7); c.setFillColor(BLACK)
        c.drawCentredString(sx+sig_w/2, y-1.1*cm, label)

    # Footer
    c.setFont('Helvetica', 6); c.setFillColor(GRAY)
    c.drawRightString(margin+w, margin+0.3*cm, 'A9.F5 Rev. 01')

    c.save(); buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f'REQ-{folio}.pdf',
                     as_attachment=False)


@app.route('/api/refacciones', methods=['GET'])
def api_get_refacciones():
    con = get_db()
    rows = con.execute('SELECT * FROM refacciones ORDER BY seccion, nombre').fetchall()
    con.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/refacciones', methods=['POST'])
def api_create_refaccion():
    d = request.json
    con = get_db()
    # Numero de parte / QR: solo se genera si el usuario eligio categoria,
    # clasificacion y proveedor (codigos del catalogo). Es opcional y no
    # afecta a las refacciones existentes, que conservan numero_parte=''.
    numero_parte = ''
    planta_cod = d.get('planta_cod', '')
    grupo = d.get('grupo', '')
    categoria_qr = d.get('categoria_qr', '')
    prov_cod = d.get('proveedor_cod', '')
    if planta_cod and grupo and categoria_qr and prov_cod:
        try:
            numero_parte = _build_numero_parte(con, planta_cod, grupo, categoria_qr, prov_cod)
        except ValueError as e:
            con.close()
            return jsonify({'error': str(e)}), 409
    con.execute("INSERT INTO refacciones (nombre,descripcion,marca,modelo,categoria,criticidad,seccion,cant_min,stock_actual,tiempo_entrega,proveedor,ubicacion,costo,notas,foto_b64,imagen_url,numero_parte) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (d.get('nombre',''),d.get('descripcion',''),d.get('marca',''),d.get('modelo',''),
         d.get('categoria',''),d.get('criticidad','MEDIA'),d.get('seccion',''),
         int(d.get('cant_min',1)),int(d.get('stock_actual',0)),
         d.get('tiempo_entrega',''),d.get('proveedor',''),d.get('ubicacion',''),
         float(d.get('costo',0)),d.get('notas',''),d.get('foto_b64',''),d.get('imagen_url',''),
         numero_parte))
    con.commit()
    new_id = con.execute('SELECT last_insert_rowid()').fetchone()[0]
    con.close()
    return jsonify({'ok': True, 'id': new_id, 'numero_parte': numero_parte})

@app.route('/api/refacciones/<int:ref_id>', methods=['PUT'])
def api_update_refaccion(ref_id):
    d = request.json
    con = get_db()
    # Patch parcial (solo foto_b64)
    if d.get('_patch'):
        fields=[]; vals=[]
        for col in ['foto_b64','imagen_url','stock_actual','cant_min']:
            if col in d: fields.append(col+'=?'); vals.append(d[col])
        if fields:
            vals.append(ref_id)
            con.execute('UPDATE refacciones SET '+','.join(fields)+',updated_at=CURRENT_TIMESTAMP WHERE id=?', vals)
            con.commit(); con.close()
        return jsonify({'ok': True})
    con.execute("UPDATE refacciones SET nombre=?,descripcion=?,marca=?,modelo=?,categoria=?,criticidad=?,seccion=?,cant_min=?,stock_actual=?,tiempo_entrega=?,proveedor=?,ubicacion=?,costo=?,notas=?,foto_b64=?,imagen_url=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (d.get('nombre',''),d.get('descripcion',''),d.get('marca',''),d.get('modelo',''),
         d.get('categoria',''),d.get('criticidad','MEDIA'),d.get('seccion',''),
         int(d.get('cant_min',1)),int(d.get('stock_actual',0)),
         d.get('tiempo_entrega',''),d.get('proveedor',''),d.get('ubicacion',''),
         float(d.get('costo',0)),d.get('notas',''),d.get('foto_b64',''),d.get('imagen_url',''),ref_id))
    con.commit()
    con.close()
    return jsonify({'ok': True})

@app.route('/api/refacciones/<int:ref_id>', methods=['DELETE'])
def api_delete_refaccion(ref_id):
    con = get_db()
    con.execute('DELETE FROM refacciones WHERE id=?', (ref_id,))
    con.commit()
    con.close()
    return jsonify({'ok': True})


# ---------------------------------------------------------------
# Generación de número de parte / QR (solo para refacciones NUEVAS)
# Basado en el catálogo de categorías/clasificaciones/proveedores
# de "base_para__QR.xlsx". No afecta refacciones ya existentes.
# ---------------------------------------------------------------

@app.route('/api/qr-catalogo', methods=['GET'])
def api_qr_catalogo():
    return jsonify({'grupos': QR_GRUPOS, 'planta': QR_PLANTA, 'proveedores': QR_PROVEEDORES})


def _build_numero_parte(con, planta_cod, grupo, categoria, prov_cod):
    """Genera un numero_parte de 8 caracteres:
       [Planta/Condicion 2][Categoria fija=8][Proveedor 2][Codigo categoria+consecutivo 3]
       Ejemplo: A08C0241 = A0 (CONDICION NORMAL) + 8 (fijo, Mantenimiento)
                + C0 (proveedor) + 241 (Electrico|Fusible=240-259, consecutivo 1)
       El consecutivo se reserva dentro del rango [desde,hasta] de la categoria,
       compartido entre TODOS los proveedores de esa categoria (asi lo definio
       generador_QR.xlsx). Se usa el primer numero libre del rango."""
    if planta_cod not in QR_PLANTA:
        raise ValueError(f"Codigo de planta/condicion invalido: {planta_cod}")
    if grupo not in QR_GRUPOS or categoria not in QR_GRUPOS[grupo]:
        raise ValueError(f"Categoria invalida: {grupo} / {categoria}")
    if prov_cod not in QR_PROVEEDORES:
        raise ValueError(f"Codigo de proveedor invalido: {prov_cod}")
    rango = QR_GRUPOS[grupo][categoria]
    desde, hasta = rango['desde'], rango['hasta']
    rows = con.execute(
        "SELECT numero_parte FROM refacciones WHERE length(numero_parte)=8"
    ).fetchall()
    usados = set()
    for r in rows:
        try:
            n = int(r['numero_parte'][-3:])
            if desde <= n <= hasta:
                usados.add(n)
        except (ValueError, TypeError, IndexError):
            continue
    for n in range(desde, hasta + 1):
        if n not in usados:
            return f"{planta_cod}{QR_CATEGORIA_FIJA}{prov_cod}{n:03d}"
    raise ValueError(f"Se agotaron los codigos disponibles para {grupo} / {categoria} ({desde}-{hasta})")


@app.route('/api/refacciones/preview-numero', methods=['GET'])
def api_preview_numero_parte():
    planta_cod = request.args.get('planta_cod', '')
    grupo = request.args.get('grupo', '')
    categoria = request.args.get('categoria', '')
    prov_cod = request.args.get('proveedor_cod', '')
    if not (planta_cod and grupo and categoria and prov_cod):
        return jsonify({'error': 'Faltan planta_cod, grupo, categoria o proveedor_cod'}), 400
    con = get_db()
    try:
        numero = _build_numero_parte(con, planta_cod, grupo, categoria, prov_cod)
    except ValueError as e:
        con.close()
        return jsonify({'error': str(e)}), 409
    con.close()
    return jsonify({'numero_parte': numero})


@app.route('/api/refacciones/export/excel', methods=['GET'])
def export_refacciones_excel():
    con = get_db()
    rows = con.execute('SELECT * FROM refacciones ORDER BY seccion, nombre').fetchall()
    con.close()

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Refacciones"

    GREEN_DARK  = PatternFill("solid", fgColor="1F5C2E")
    GREEN_MED   = PatternFill("solid", fgColor="2E7D32")
    GREEN_LIGHT = PatternFill("solid", fgColor="C8E6C9")
    GRAY_LIGHT  = PatternFill("solid", fgColor="F5F5F5")
    RED_FILL    = PatternFill("solid", fgColor="C62828")
    ORG_FILL    = PatternFill("solid", fgColor="E65100")
    GREEN_FILL  = PatternFill("solid", fgColor="2E7D32")
    thin = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:M1")
    ws["A1"] = "LISTA DE REFACCIONES  —  ALMACÉN MANTENIMIENTO  |  AD-PACK  Termoformado"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = GREEN_DARK
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle row
    ws.merge_cells("A2:D2"); ws["A2"] = f"Área:  Mantenimiento / Termoformado"
    ws.merge_cells("E2:H2"); ws["E2"] = "Responsable:  _________________"
    ws.merge_cells("I2:J2"); ws["I2"] = "Revisión:  00"
    ws.merge_cells("K2:M2"); ws["K2"] = f"Fecha:  {date.today()}"
    for cell in ["A2","E2","I2","K2"]:
        ws[cell].font = Font(bold=True, size=9, color="FFFFFF")
        ws[cell].fill = GREEN_MED
        ws[cell].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 18

    # Header row
    headers = ["#","REFACCION / DESCRIPCION","MARCA","MODELO","CATEGORIA","CRITICIDAD","SECCION","CANT.MIN.","STOCK ACT.","T. ENTREGA","PROVEEDOR","UBICACION ALMACEN"]
    col_w   = [5, 38, 18, 16, 14, 12, 20, 7, 7, 12, 22, 22]
    for i, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = GREEN_MED
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 30

    # Group by section
    sections = {}
    for r in rows:
        sec = r["seccion"] or "Sin sección"
        sections.setdefault(sec, []).append(r)

    row_num = 4
    num = 1
    for sec, items in sections.items():
        # Section header
        ws.merge_cells(f"A{row_num}:M{row_num}")
        ws[f"A{row_num}"] = f"  ▌ {sec.upper()}"
        ws[f"A{row_num}"].font = Font(bold=True, color="FFFFFF", size=10)
        ws[f"A{row_num}"].fill = GREEN_MED
        ws[f"A{row_num}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        for r in items:
            fill = GRAY_LIGHT if num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            crit = r["criticidad"] or "MEDIA"
            crit_fill = RED_FILL if crit=="ALTA" else (ORG_FILL if crit=="MEDIA" else GREEN_FILL)

            vals = [num, r["nombre"], r["marca"] or "", r["modelo"] or "",
                    r["categoria"] or "", crit, r["seccion"] or "",
                    r["cant_min"], r["stock_actual"],
                    r["tiempo_entrega"] or "", r["proveedor"] or "",
                    r["ubicacion"] or ""]

            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = border
                cell.font = Font(size=9)
                cell.alignment = Alignment(vertical="center", wrap_text=True,
                                           horizontal="center" if col in [1,8,9] else "left")
                if col == 6:
                    cell.fill = crit_fill
                    cell.font = Font(bold=True, color="FFFFFF", size=9)
                else:
                    cell.fill = fill

            # Stock bajo warning
            if r["stock_actual"] <= r["cant_min"]:
                ws.cell(row=row_num, column=9).font = Font(bold=True, color="C62828", size=9)

            ws.row_dimensions[row_num].height = 16
            row_num += 1
            num += 1

    # Footer
    ws.merge_cells(f"A{row_num}:M{row_num}")
    ws[f"A{row_num}"] = "  🔴 ALTA = Paro de producción    |    🟡 MEDIA = Afecta eficiencia    |    🟢 BAJA = Preventivo"
    ws[f"A{row_num}"].font = Font(italic=True, size=8, color="555555")
    ws[f"A{row_num}"].fill = GREEN_LIGHT

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"Lista_Refacciones_{date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/api/refacciones/export/pdf', methods=['GET'])
def export_refacciones_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from datetime import date

    con = get_db()
    rows = con.execute('SELECT * FROM refacciones ORDER BY seccion, nombre').fetchall()
    con.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=8*mm, rightMargin=8*mm,
                            topMargin=8*mm, bottomMargin=8*mm)

    GREEN_DARK  = colors.HexColor('#1F5C2E')
    GREEN_MED   = colors.HexColor('#2E7D32')
    GREEN_LIGHT = colors.HexColor('#C8E6C9')
    RED   = colors.HexColor('#C62828')
    ORG   = colors.HexColor('#E65100')
    GRAY  = colors.HexColor('#F5F5F5')
    WHITE = colors.white

    def ps(size=7, bold=False, color=colors.black, align=TA_LEFT):
        return ParagraphStyle('x', fontSize=size, leading=size+2,
                              fontName='Helvetica-Bold' if bold else 'Helvetica',
                              textColor=color, alignment=align)

    crit_col = {'ALTA': RED, 'MEDIA': ORG, 'BAJA': GREEN_MED}
    CW = [7*mm,65*mm,22*mm,18*mm,16*mm,16*mm,10*mm,10*mm,14*mm,28*mm,30*mm]

    sections = {}
    for r in rows:
        sections.setdefault(r['seccion'] or 'Sin seccion', []).append(r)

    story = []

    # ---- Titulo ----
    t = Table([[Paragraph('LISTA DE REFACCIONES — ALMACÉN MANTENIMIENTO | AD-PACK Termoformado',
                          ps(12, True, WHITE, TA_CENTER))]],
              colWidths=[sum(CW)])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),GREEN_DARK),
                           ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
    story.append(t)

    s = Table([[f'Área: Mantenimiento / Termoformado',
                'Responsable: _________________',
                f'Revisión: 00   |   Fecha: {date.today()}']],
              colWidths=[sum(CW)//3, sum(CW)//3, sum(CW)-2*(sum(CW)//3)])
    s.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),GREEN_MED),
                           ('TEXTCOLOR',(0,0),(-1,-1),WHITE),
                           ('FONTSIZE',(0,0),(-1,-1),7.5),
                           ('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),
                           ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.append(s)
    story.append(Spacer(1,2*mm))

    num = 1
    HDRS = ['#','REFACCION / DESCRIPCION','MARCA','MODELO','CATEGORIA','CRITICIDAD',
            'MIN','STOCK','ENTREGA','PROVEEDOR','UBICACION ALMACEN']

    for sec, items in sections.items():
        sh = Table([[Paragraph(f'  ▌ {sec.upper()}', ps(8, True, WHITE))]], colWidths=[sum(CW)])
        sh.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),GREEN_MED),
                                ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
        story.append(sh)

        tdata = [[Paragraph(h, ps(7,True,WHITE,TA_CENTER)) for h in HDRS]]
        tstyle = [
            ('BACKGROUND',(0,0),(-1,0),GREEN_MED),
            ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#BDBDBD')),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
        ]

        for i, r in enumerate(items, 1):
            low = r['stock_actual'] <= r['cant_min']
            cc  = crit_col.get(r['criticidad'] or 'MEDIA', ORG)
            bg  = GRAY if i % 2 == 0 else WHITE
            tdata.append([
                Paragraph(str(num),           ps(7,align=TA_CENTER)),
                Paragraph(str(r['nombre'] or ''),     ps(7)),
                Paragraph(str(r['marca'] or ''),      ps(7)),
                Paragraph(str(r['modelo'] or ''),     ps(7)),
                Paragraph(str(r['categoria'] or ''),  ps(7)),
                Paragraph(str(r['criticidad'] or ''), ps(7,True,WHITE,TA_CENTER)),
                Paragraph(str(r['cant_min']),          ps(7,align=TA_CENTER)),
                Paragraph(str(r['stock_actual']),      ps(7,bold=low,color=RED if low else colors.black,align=TA_CENTER)),
                Paragraph(str(r['tiempo_entrega'] or ''), ps(7)),
                Paragraph(str(r['proveedor'] or ''),   ps(7)),
                Paragraph(str(r['ubicacion'] or ''),   ps(7)),
            ])
            tstyle += [('BACKGROUND',(0,i),(-1,i),bg),
                       ('BACKGROUND',(5,i),(5,i),cc)]
            num += 1

        dt = Table(tdata, colWidths=CW, repeatRows=1)
        dt.setStyle(TableStyle(tstyle))
        story.append(dt)
        story.append(Spacer(1,2*mm))

    # Footer
    ft = Table([[Paragraph('🔴 ALTA = Paro producción  |  🟡 MEDIA = Afecta eficiencia  |  🟢 BAJA = Preventivo  |  Stock rojo = bajo mínimo',
                            ps(6, color=colors.HexColor('#555555')))]],
               colWidths=[sum(CW)])
    ft.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),GREEN_LIGHT),
                            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.append(ft)

    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"Lista_Refacciones_{date.today()}.pdf",
                     mimetype='application/pdf')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 3000)), debug=False)
